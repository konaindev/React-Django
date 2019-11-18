import os

from openpyxl import Workbook
from openpyxl.styles import PatternFill

from remark.projects.models import Project, Period
from remark.projects.reports.periods import ComputedPeriod


def get_periods(project):
    periods_q = Period.objects.filter(project=project).order_by("start")
    periods = []
    for p in periods_q:
        if (p.end - p.start).days == 7:
            periods.append(ComputedPeriod(p))
    return periods


def get_trends_value(item, trend, periods, i, project_name):
    i_start = i - len(trend) + 1
    i_end = i_start + 15
    return {
        "project_name": project_name,
        "start_downtrend": trend[0].start,
        "end_downtrend": trend[-1].end,
        "number_weeks": len(trend),
        "kpi1_before": getattr(trend[0], item["kpi1"]),
        "kpi1_after": getattr(trend[-1], item["kpi1"]),
        "kpi2_before": getattr(trend[0], item["kpi2"]),
        "kpi2": [getattr(p, item["kpi2"]) for p in periods[i_start:i_end]],
    }


def research(projects, item):
    result = []

    for project in projects:
        periods = get_periods(project)
        trend = []
        for i, p in enumerate(periods):
            if len(trend) == 0:
                trend.append(p)
            else:
                kpi_value = getattr(p, item["kpi1"])
                before_kpi_value = getattr(trend[-1], item["kpi1"])
                if before_kpi_value >= kpi_value:
                    trend.append(p)
                else:
                    if len(trend) != 1:
                        result.append(
                            get_trends_value(item, trend, periods, i, project.name)
                        )
                    trend = [p]
        if len(trend) > 1:
            result.append(
                get_trends_value(item, trend, periods, len(periods), project.name)
            )
    return result


def create_xls(kpis, kpis_data):
    FILE_NAME = os.path.join(
        os.path.abspath(os.path.dirname(__file__)), "leading_indicator.xlsx"
    )

    wb = Workbook()
    ws = wb.active

    for i, kpi in enumerate(kpis):
        ws.title = kpi["kpi1_name"]
        ws.cell(column=1, row=1, value="Project Name")
        ws.cell(column=2, row=1, value="Start of downtrend")
        ws.cell(column=3, row=1, value="End of downtrend")
        ws.cell(column=4, row=1, value="Number Weeks of downtrend")
        ws.cell(column=5, row=1, value=f"{kpi['kpi1_name']} Before Down Trend")
        ws.cell(column=6, row=1, value=f"{kpi['kpi1_name']} Last Week of Down Trend")
        ws.cell(column=7, row=1, value=f"{kpi['kpi2_name']} Before Trend")
        for column in range(1, 16):
            ws.cell(column=7 + column, row=1, value=f"{kpi['kpi2_name']} Week {column}")

        row = 1
        for data in kpis_data[i]:
            row += 1
            ws.cell(column=1, row=row, value=data["project_name"])
            ws.cell(column=2, row=row, value=data["start_downtrend"])
            ws.cell(column=3, row=row, value=data["end_downtrend"])
            ws.cell(column=4, row=row, value=data["number_weeks"])
            ws.cell(column=5, row=row, value=data["kpi1_before"])
            ws.cell(column=6, row=row, value=data["kpi1_after"])
            ws.cell(column=7, row=row, value=data["kpi2_before"])
            before = data["kpi2_before"]
            for j, v in enumerate(data["kpi2"]):
                if before < v:
                    bg_color = "66FF66"
                elif before == v:
                    bg_color = "FFFF66"
                else:
                    bg_color = "FC2847"
                column = j + 8
                before = v
                ws.cell(column=column, row=row, value=v).fill = PatternFill(
                    fgColor=bg_color, fill_type="solid"
                )

        ws = wb.create_sheet()

    wb.save(FILE_NAME)


def run():
    kpis = [
        {
            "kpi1_name": "USV",
            "kpi1": "usvs",
            "kpi2": "lease_applications",
            "kpi2_name": "Lease Applications",
        },
        {
            "kpi1_name": "INQ",
            "kpi1": "inquiries",
            "kpi2": "lease_applications",
            "kpi2_name": "Lease Applications",
        },
        {
            "kpi1_name": "INQ>TOU",
            "kpi1": "inq_tou_perc",
            "kpi2": "lease_applications",
            "kpi2_name": "Lease Applications",
        },
        {
            "kpi1_name": "TOU>APP",
            "kpi1": "tou_app_perc",
            "kpi2": "lease_applications",
            "kpi2_name": "Lease Applications",
        },
        {
            "kpi1_name": "APP",
            "kpi1": "lease_applications",
            "kpi2": "leased_rate",
            "kpi2_name": "Leased Rate",
        },
    ]

    result = []
    projects = Project.objects.all()
    for kpi in kpis:
        result.append(research(projects, kpi))

    create_xls(kpis, result)
