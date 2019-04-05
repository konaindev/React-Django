"""
Utilities for validating and importing spreadsheet uploads.
"""
from collections import namedtuple
import re

import openpyxl

from remark.lib.math import d_quant_currency
from .models import Spreadsheet


_parse_re = re.compile("(?:'?([a-z 0-9_-]+)'?!)?([a-z]*)([0-9]*)", re.IGNORECASE)


def parse_location(location):
    """
    Attempt to parse a location string. 
    
    Return a tuple of (sheet, col, row) from the parse, where one
    or more of those values could be None.

    Any of the following are valid strings:

        "sheetname!"
        "sheetname!C3"
        "'sheet name'!C3"
        "sheetname!C'
        "sheetname!3'
        "C3"
        "C"
        "3"
    """
    sheet, col, row = list(_parse_re.match(location).groups())
    return (sheet or None, col.upper() if col else None, int(row) if row else None)


def parse_location_or_default(location=None, sheet=None, col=None, row=None):
    """
    Parse a location string, falling back to defaults it the location string
    does not explicitly provide them.
    """
    _sheet, _col, _row = parse_location(location) if location else (None, None, None)
    return (_sheet or sheet, _col or col, _row or row)


def unparse_location(sheet=None, col=None, row=None):
    return f"'{sheet}'!{col}{row}"


class DataType:
    STRING = openpyxl.cell.cell.TYPE_STRING
    FORMULA = openpyxl.cell.cell.TYPE_FORMULA
    NUMERIC = openpyxl.cell.cell.TYPE_NUMERIC
    BOOL = openpyxl.cell.cell.TYPE_BOOL
    NULL = openpyxl.cell.cell.TYPE_NULL
    ERROR = openpyxl.cell.cell.TYPE_ERROR
    DATETIME = "d"  # This requires us to call is_date()


class ExcelError(Exception):
    """Generic error for all code in this library."""

    def __init__(self, where=None, message=None):
        """
        Construct an ExcelError. For convenience, `where` may be one of:

        1. An arbitrary string
        2. A tuple of (sheet, col, row)
        3. An openpyxl Cell
        """
        if isinstance(where, openpyxl.cell.cell.Cell):
            where = unparse_location(where.parent.title, where.column_name, where.row)
        elif isinstance(where, tuple):
            where = unparse_location(*where)

        super().__init__(f"{where}:: {message}")


class ExcelValidationError(ExcelError):
    """An exception for when there's an error in the excel spreadsheet itself."""

    pass


class ExcelProgrammingError(ExcelError):
    """An exception for when we messed something up here in this code..."""

    pass


def row_range(start_row, end_row):
    """Return an iterable of rows; both start and end are inclusive."""
    return range(start_row, end_row + 1)


def index_for_col(col):
    """Given a column name, return an index for it; A = 1"""
    # This is simple base-26 arithmetic, only there's no 0 in the alphabet
    # We reverse the column name because the rightmost character is the 0's
    # place; the next rightmost is the 26's place, etc.
    return sum(((ord(c) - ord("A") + 1) * 26 ** i for i, c in enumerate(col[::-1])))


def col_for_index(index):
    """Given a column index with A = 1, return a name for it."""
    # This is base-10 to base-26 arithmetic, shifted by 1
    col = ""
    while index > 0:
        index, r = divmod(index - 1, 26)
        col = chr(ord("A") + r) + col
    return col


def next_col(col):
    """Given a column name, determine the next column name."""
    return col_for_index(index_for_col(col) + 1)


def col_range(start_col, end_col):
    """Return an iterable of cols; both start and end are inclusive."""
    return (
        col_for_index(i)
        for i in range(index_for_col(start_col), index_for_col(end_col) + 1)
    )


SchemaItem = namedtuple("SchemaItem", ["getter", "data_type", "converter"])


def base_getter(workbook, sheet, col, row):
    if workbook is None:
        raise ExcelProgrammingError(message="No workbook found")
    if (not sheet) or (not col) or (not row):
        raise ExcelProgrammingError((sheet, col, row), "incomplete location")
    if sheet not in workbook:
        raise ExcelValidationError(
            (sheet, col, row), f"'{sheet}' not found in workbook"
        )
    return workbook[sheet][f"{col}{row}"]


def loc(location=None, sheet=None, col=None, row=None):
    """
    Returns a cell getter function that is pre-populated with an arbitrary set
    of values.
    """
    d_sheet, d_col, d_row = parse_location_or_default(location, sheet, col, row)

    def getter(workbook, sheet, col, row):
        return base_getter(workbook, sheet or d_sheet, col or d_col, row or d_row)

    return getter


MATCHERS = {
    "eq": lambda v, t: v == t,
    "gt": lambda v, t: v > t,
    "gte": lambda v, t: v >= t,
    "lt": lambda v, t: v < t,
    "lte": lambda v, t: v <= t,
    "exact": lambda v, t: v == t,  # synonym for eq
    "iexact": lambda v, t: v.casefold() == t.casefold(),
    "startswith": lambda v, t: v.startswith(t),
    "istartswith": lambda v, t: v.casefold().startswith(t.casefold()),
    "contains": lambda v, t: t in v,
    "icontains": lambda v, t: t.casefold() in v.casefold(),
    "endswith": lambda v, t: v.endswith(t),
    "iendswith": lambda v, t: v.casefold().endswith(t.casefold()),
}


def match(v, **query):
    """
    Returns True if the `v` matches the constraints provided in `query`.

    The `query` parameters provide a set of simple but flexible matching options,
    including:

        eq, gt, gte, lt, lte, 
        exact, iexact, startswith, istartswith, contains, icontains, endswith, iendswith

    These are similar to value filters in Django querysets.
    """
    return all((MATCHERS[k](v, t) for k, t in query.items()))


def matchp(**query):
    """
    Returns a predicate matcher method.
    """
    return lambda v: match(v, **query)


def imatchp(t):
    """
    Convenience; return a predicate matcher that always uses icontains.
    """
    return matchp(icontains=t)


def find_col(header_row, predicate, start_col="A", end_col="ZZ"):
    """
    Returns a getter method that locates a cell by looking for the column
    that matches a given `predicate` in the provided `header_row`.

    As a convenience, if `predicate` is a string, a match method is created
    using an `icontains=predicate` query. That's the form of query you
    probably want most of the time.

    The `header_row` is scanned from `start_col` to `end_col`, which default
    to "reasonable" values.
    """
    if isinstance(predicate, str):
        predicate = matchp(icontains=predicate)

    # We only want to perform the search once, so we cache the found
    # column in the outer scope.
    cached_col = None

    def find(workbook, sheet):
        """Perform the actual search."""
        # Gin up a sequence of columns whose value in the header_row
        # satisfy the predicate
        seq = (
            col
            for col in col_range(start_col, end_col)
            if predicate(base_getter(workbook, sheet, col, header_row).value)
        )
        # Return the first item in the sequence, or None
        return next(seq, None)

    def getter(workbook, sheet, col, row):
        nonlocal cached_col
        if cached_col is None:
            cached_col = find(workbook, sheet)
        return base_getter(workbook, sheet, cached_col or col, row)

    return getter


def find_row(header_col, predicate, start_row=1, end_row=100):
    """
    Returns a getter method that locates a cell by looking for the row
    that matches a given `predicate` in the provided `header_col`.

    As a convenience, if `predicate` is a string, a match method is created
    using an `icontains=predicate` query. That's the form of query you
    probably want most of the time.

    The `header_col` is scanned from `start_row` to `end_row`, which default
    to "reasonable" values.
    """
    if isinstance(predicate, str):
        predicate = matchp(icontains=predicate)

    # We only want to perform the search once, so we cache the found
    # row in the outer scope.
    cached_row = None

    def find(workbook, sheet):
        """Perform the actual search."""
        # Gin up a sequence of rows whose value in the header_col
        # satisfy the predicate
        seq = (
            row
            for row in col_range(start_row, end_row)
            if predicate(base_getter(workbook, sheet, header_col, row).value)
        )
        # Return the first item in the sequence, or None
        return next(seq, None)

    def getter(workbook, sheet, col, row):
        nonlocal cached_row
        if cached_row is None:
            cached_row = find(workbook, sheet)
        return base_getter(workbook, sheet, col, cached_row or row)

    return getter


class ExcelImporter:
    """
    Provides bare-bones tools for importing *any* of our data entry spreadsheets.
    """

    def __init__(self, f):
        """
        Create an importer with a fileobj or with the path to an extant file.
        """
        self.workbook = openpyxl.load_workbook(f, data_only=True, read_only=True)
        self.cleaned_data = {}
        self.errors = []

    def check_data_type(self, cell, data_type):
        """Raise an exception if cell's data type doesn't match provided data_type."""
        # If we have no expectations, we're always happy
        if data_type == DataType.DATETIME:
            valid = cell.is_date
        else:
            valid = data_type == cell.data_type
        if not valid:
            raise ExcelValidationError(
                cell, f"found data type '{cell.data_type}' but expected '{data_type}'"
            )

    def check_convert(self, cell, converter=None):
        """
        Attempt to convert a cell's raw value to a python value.

        Raise an exception if the conversion fails. Otherwise, return
        the converted value.
        """
        # CONSIDER this is a pretty awkward API, I think?
        value = cell.value
        if converter is not None:
            try:
                value = converter(value)
            except Exception as e:
                raise ExcelValidationError(
                    cell, f"Could not apply value converter '{converter}': '{e}'"
                )
        return value

    def cell(self, getter, sheet=None, col=None, row=None):
        """
        Given a getter function, return the cell at the underlying location.
        """
        return getter(self.workbook, sheet, col, row)

    def schema_value(self, schema_item, sheet=None, col=None, row=None):
        """
        Given a SchemaItem, return the python-converted value at the
        underlying location, validating it as necessary.
        """
        cell = self.cell(schema_item.getter, sheet, col, row)
        self.check_data_type(cell, schema_item.data_type)
        value = self.check_convert(cell, schema_item.converter)
        return value

    def check_schema_value(
        self, schema_item, sheet=None, col=None, row=None, expected=None
    ):
        """
        Raise an exception if the wrong value is found.

        `expected` can be an explicit value, or a callable; if a callable
        is provided, it is called with the converted value and must return
        True if the expectation is met.
        """
        cell = self.cell(schema_item.getter, sheet, col, row)
        self.check_data_type(cell, schema_item.data_type)
        value = self.check_convert(cell, schema_item.converter)
        value_matches = expected(value) if callable(expected) else expected == value
        if not value_matches:
            raise ExcelValidationError(
                cell, f"expected value '{expected}', found '{value}'"
            )

    def row(self, schema, row, sheet=None):
        """
        Return the structured contents of a given row, based on the provided
        schema definition.

        A schema definition is simply a dictionary mapping a key name to
        a SchemaItem, which provides a getter, an expcted data type, and
        a python type converter.
        """
        return {
            key: self.schema_value(schema_item, sheet=sheet, row=row)
            for key, schema_item in schema.items()
        }

    def col(self, schema, col, sheet=None):
        """
        Return the structured contents of a given column, based on the
        provided schema definition.

        A schema definition is simply a dictionary mapping a key name to
        a SchemaItem, which provides a getter, an expcted data type, and
        a python type converter.
        """
        return {
            key: self.schema_value(schema_item, sheet=sheet, col=col)
            for key, schema_item in schema.items()
        }

    def row_table(self, schema, start_row, end_row, sheet=None):
        """
        Return an array of rows, starting with start_row and including
        end_row.
        """
        return [
            self.row(schema, row, sheet=sheet) for row in row_range(start_row, end_row)
        ]

    def col_table(self, schema, start_col, end_col, sheet=None):
        """
        Return an array of columns, starting with start_col and including
        end_col.
        """
        return [
            self.col(schema, col, sheet=sheet) for col in col_range(start_col, end_col)
        ]

    def is_valid(self):
        """Validate the spreadsheet; return False if not possible."""
        try:
            self.clean()
        except ExcelValidationError as e:
            self.errors.append(e)
        return len(self.errors) == 0

    def clean(self):
        """
        Validate the spreadsheet.

        Updates the underlying cleaned_data dictionary as appropriate.

        Adds ExcelValidationErrors to the errors list when encountering
        validation errors that do not impact the ability to further validate;
        raises ExcelValidationError when encountering a hard-stop validation
        error.
        """
        pass


class RemarkablyExcelImporter(ExcelImporter):
    """
    An excel importer base class with tools specifically tailored to
    remarkably-style excel templates.
    """

    expected_type = None
    expected_version = None

    SPREADSHEET_TYPE_SCHEMA_ITEM = SchemaItem(loc("VERSION!B1"), DataType.STRING, str)
    SPREADSHEET_VERSION_SCHEMA_ITEM = SchemaItem(
        loc("VERSION!B2"), DataType.NUMERIC, int
    )

    def clean(self):
        """
        Clean a remarkably-style excel spreadsheet.
        """
        self.check_version()

    def check_version(self):
        """
        Validate the VERSION tab of the spreadsheet.
        """
        self.check_schema_value(
            self.SPREADSHEET_TYPE_SCHEMA_ITEM, expected=self.expected_type
        )
        self.check_schema_value(
            self.SPREADSHEET_VERSION_SCHEMA_ITEM, expected=self.expected_version
        )


def date_from_datetime(d):
    return d.date()


class BaselinePerfImporter(RemarkablyExcelImporter):
    expected_type = "baseline_perf"
    expected_version = 1

    DATES_VALID = SchemaItem(loc("META!B11"), DataType.STRING, str)
    BASELINE_PERIODS = SchemaItem(loc("META!B5"), DataType.NUMERIC, int)
    START_ROW = SchemaItem(loc("META!B1"), DataType.NUMERIC, int)
    END_ROW = SchemaItem(loc("META!B4"), DataType.NUMERIC, int)
    BASELINE_START_DATE = SchemaItem(
        loc("META!B7"), DataType.DATETIME, date_from_datetime
    )
    BASELINE_END_DATE = SchemaItem(
        loc("META!B8"), DataType.DATETIME, date_from_datetime
    )

    HEADER_ROW = 2

    PERIOD_SHEET = "output_periods"

    PERIOD_ROW_SCHEMA = {
        "start": SchemaItem(
            find_col(HEADER_ROW, "start date"), DataType.DATETIME, date_from_datetime
        ),
        "end": SchemaItem(
            find_col(HEADER_ROW, "end date"), DataType.DATETIME, date_from_datetime
        ),
        "leased_units_start": SchemaItem(
            find_col(HEADER_ROW, "leased units @ start"), DataType.NUMERIC, int
        ),
        "leases_ended": SchemaItem(
            find_col(HEADER_ROW, "ended"), DataType.NUMERIC, int
        ),
        "lease_applications": SchemaItem(
            find_col(HEADER_ROW, "APPs"), DataType.NUMERIC, int
        ),
        "leases_executed": SchemaItem(
            find_col(HEADER_ROW, "EXEs"), DataType.NUMERIC, int
        ),
        "lease_cds": SchemaItem(find_col(HEADER_ROW, "CDs"), DataType.NUMERIC, int),
        "lease_renewal_notices": SchemaItem(
            find_col(HEADER_ROW, "Notices: Renewals"), DataType.NUMERIC, int
        ),
        "lease_renewals": SchemaItem(
            # Use matchp(iexact=...) to disambiguate with "Notices: Renewals"
            find_col(HEADER_ROW, matchp(iexact="Renewals")),
            DataType.NUMERIC,
            int,
        ),
        "lease_vacation_notices": SchemaItem(
            find_col(HEADER_ROW, "Notices: Vacate"), DataType.NUMERIC, int
        ),
        "occupiable_units_start": SchemaItem(
            find_col(HEADER_ROW, "occupiable units"), DataType.NUMERIC, int
        ),
        "occupied_units_start": SchemaItem(
            find_col(HEADER_ROW, "occupied units"), DataType.NUMERIC, int
        ),
        "move_ins": SchemaItem(find_col(HEADER_ROW, "move ins"), DataType.NUMERIC, int),
        "move_outs": SchemaItem(
            find_col(HEADER_ROW, "move outs"), DataType.NUMERIC, int
        ),
        "acq_reputation_building": SchemaItem(
            find_col(HEADER_ROW, "Reputation ACQ"), DataType.NUMERIC, int
        ),
        "acq_demand_creation": SchemaItem(
            find_col(HEADER_ROW, "Demand ACQ"), DataType.NUMERIC, d_quant_currency
        ),
        "acq_leasing_enablement": SchemaItem(
            find_col(HEADER_ROW, "Leasing ACQ"), DataType.NUMERIC, d_quant_currency
        ),
        "acq_market_intelligence": SchemaItem(
            find_col(HEADER_ROW, "Market ACQ"), DataType.NUMERIC, d_quant_currency
        ),
        "ret_reputation_building": SchemaItem(
            find_col(HEADER_ROW, "Reputation RET"), DataType.NUMERIC, d_quant_currency
        ),
        "ret_demand_creation": SchemaItem(
            find_col(HEADER_ROW, "Demand RET"), DataType.NUMERIC, d_quant_currency
        ),
        "ret_leasing_enablement": SchemaItem(
            find_col(HEADER_ROW, "Leasing RET"), DataType.NUMERIC, d_quant_currency
        ),
        "ret_market_intelligence": SchemaItem(
            find_col(HEADER_ROW, "Market RET"), DataType.NUMERIC, d_quant_currency
        ),
        "usvs": SchemaItem(find_col(HEADER_ROW, "USVs"), DataType.NUMERIC, int),
        "inquiries": SchemaItem(find_col(HEADER_ROW, "INQs"), DataType.NUMERIC, int),
        "tours": SchemaItem(find_col(HEADER_ROW, "TOUs"), DataType.NUMERIC, int),
    }

    def check_meta(self):
        """
        Validate that the basic contents of our META tab are valid.
        """
        self.check_schema_value(self.DATES_VALID, expected="valid")
        self.check_schema_value(self.BASELINE_PERIODS, expected=lambda value: value > 0)

    def clean(self):
        super().clean()
        self.check_meta()
        start_row = self.schema_value(self.START_ROW)
        end_row = self.schema_value(self.END_ROW)
        self.cleaned_data["baseline_start_date"] = self.schema_value(
            self.BASELINE_START_DATE
        )
        self.cleaned_data["baseline_end_date"] = self.schema_value(
            self.BASELINE_END_DATE
        )
        self.cleaned_data["periods"] = self.row_table(
            schema=self.PERIOD_ROW_SCHEMA,
            start_row=start_row,
            end_row=end_row,
            sheet=self.PERIOD_SHEET,
        )


# TODO where should this go? -Dave
IMPORTERS_FOR_KIND = {Spreadsheet.KIND_PERIODS: BaselinePerfImporter}


def get_importer(kind, f):
    importer_class = IMPORTERS_FOR_KIND.get(kind)
    return importer_class(f) if importer_class is not None else None
