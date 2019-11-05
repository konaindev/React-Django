import datetime
import decimal
import os.path
import json

from django.contrib.auth.models import Group
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook
from io import BytesIO
from unittest import mock

from remark.crm.models import Business
from remark.geo.mocks import mocked_geocode
from remark.geo.models import Address
from remark.users.models import Account, User
from remark.lib.metrics import BareMultiPeriod
from remark.email_app.invites.added_to_property import (
    send_invite_email,
    get_template_vars,
)

from .models import Building, Fund, LeaseStage, Period, Project, Property, TargetPeriod
from .reports.periods import ComputedPeriod
from .reports.performance import PerformanceReport
from .export import export_periods_to_csv, export_periods_to_excel
from remark.settings import BASE_URL


def create_project(project_name="project 1"):
    address = Address.objects.create(
        street_address_1="2284 W. Commodore Way, Suite 200",
        city="Seattle",
        state="WA",
        zip_code=98199,
        country="US",
    )
    account = Account.objects.create(
        company_name="test", address=address, account_type=4
    )
    asset_manager = Business.objects.create(
        name="Test Asset Manager", is_asset_manager=True
    )
    property_manager = Business.objects.create(
        name="Test Property Manager", is_property_manager=True
    )
    property_owner = Business.objects.create(
        name="Test Property Owner", is_property_owner=True
    )
    fund = Fund.objects.create(account=account, name="Test Fund")
    property = Property.objects.create(
        name="property 1",
        average_monthly_rent=decimal.Decimal("0"),
        lowest_monthly_rent=decimal.Decimal("0"),
        geo_address=address,
    )
    group = Group.objects.create(name=f"{project_name} view group")
    project = Project.objects.create(
        name=project_name,
        baseline_start=datetime.date(year=2018, month=11, day=19),
        baseline_end=datetime.date(year=2018, month=12, day=26),
        account=account,
        asset_manager=asset_manager,
        property_manager=property_manager,
        property_owner=property_owner,
        fund=fund,
        property=property,
        view_group=group,
    )
    return project, group


def add_user_to_group(group, email="test@remarkably.io"):
    user, _ = User.objects.get_or_create(email=email)
    group.user_set.add(user)


class DefaultComputedPeriodTestCase(TestCase):
    """
    Test that all computed properties on a default Period instance
    return sane values.
    """

    def setUp(self):
        project, _ = create_project()
        stage = LeaseStage.objects.get(short_name="performance")
        raw_period = Period.objects.create(
            project=project,
            lease_stage=stage,
            start=datetime.date(year=2018, month=12, day=19),
            end=datetime.date(year=2018, month=12, day=26),
            leased_units_start=0,
            occupiable_units_start=0,
            occupied_units_start=0,
        )
        raw_target_period = TargetPeriod.objects.create(
            project=project,
            start=datetime.date(year=2018, month=12, day=19),
            end=datetime.date(year=2018, month=12, day=26),
        )
        multiperiod = BareMultiPeriod.from_periods([raw_period, raw_target_period])
        period = multiperiod.get_cumulative_period()
        self.period = ComputedPeriod(period)

    def test_delta_leases(self):
        self.assertEqual(self.period.delta_leases, 0)

    def test_leased_units(self):
        self.assertEqual(self.period.leased_units, 0)

    def test_leased_rate(self):
        self.assertEqual(self.period.leased_rate, 0)

    def test_renewal_rate(self):
        self.assertEqual(self.period.renewal_rate, 0)

    def test_lease_cd_rate(self):
        self.assertEqual(self.period.lease_cd_rate, 0)

    def test_target_leased_units(self):
        self.assertEqual(self.period.target_leased_units, None)

    def test_occupied_units(self):
        self.assertEqual(self.period.occupied_units, 0)

    def test_occupancy_rate(self):
        self.assertEqual(self.period.occupancy_rate, 0)

    def test_acq_investment(self):
        self.assertEqual(self.period.acq_investment, 0)

    def test_ret_investment(self):
        self.assertEqual(self.period.ret_investment, 0)

    def test_investment(self):
        self.assertEqual(self.period.investment, 0)

    def test_estimated_acq_revenue_gain(self):
        self.assertEqual(self.period.estimated_acq_revenue_gain, 0)

    def test_estimated_ret_revenue_gain(self):
        self.assertEqual(self.period.estimated_ret_revenue_gain, 0)

    def test_acq_romi(self):
        self.assertEqual(self.period.acq_romi, 0)

    def test_ret_romi(self):
        self.assertEqual(self.period.ret_romi, 0)

    def test_romi(self):
        self.assertEqual(self.period.romi, 0)

    def test_target_investment(self):
        self.assertEqual(self.period.target_investment, None)

    def test_target_estimated_acq_revenue_gain(self):
        self.assertEqual(self.period.target_estimated_acq_revenue_gain, None)

    def test_target_estimated_ret_revenue_gain(self):
        self.assertEqual(self.period.target_estimated_ret_revenue_gain, None)

    def test_target_acq_romi(self):
        self.assertEqual(self.period.target_acq_romi, None)

    def test_target_ret_romi(self):
        self.assertEqual(self.period.target_ret_romi, None)

    def test_target_romi(self):
        self.assertEqual(self.period.target_romi, None)

    def test_usv_inq_perc(self):
        self.assertEqual(self.period.usv_inq_perc, 0)

    def test_inq_tou_perc(self):
        self.assertEqual(self.period.inq_tou_perc, 0)

    def test_tou_app_perc(self):
        self.assertEqual(self.period.tou_app_perc, 0)

    def test_app_exe_perc(self):
        self.assertEqual(self.period.app_exe_perc, 0)

    def test_usv_exe_perc(self):
        self.assertEqual(self.period.usv_exe_perc, 0)

    def test_target_usv_inq_perc(self):
        self.assertEqual(self.period.target_usv_inq_perc, None)

    def test_target_inq_tou_perc(self):
        self.assertEqual(self.period.target_inq_tou_perc, None)

    def test_target_tou_app_perc(self):
        self.assertEqual(self.period.target_tou_app_perc, None)

    def test_target_app_exe_perc(self):
        self.assertEqual(self.period.target_app_exe_perc, None)

    def test_target_usv_exe_perc(self):
        self.assertEqual(self.period.target_usv_exe_perc, None)

    def test_cost_per_usv(self):
        self.assertEqual(self.period.cost_per_usv, 0)

    def test_cost_per_inq(self):
        self.assertEqual(self.period.cost_per_inq, 0)

    def test_cost_per_tou(self):
        self.assertEqual(self.period.cost_per_tou, 0)

    def test_cost_per_app(self):
        self.assertEqual(self.period.cost_per_app, 0)

    def test_cost_per_exe(self):
        self.assertEqual(self.period.cost_per_exe, 0)

    def test_target_cost_per_usv(self):
        self.assertEqual(self.period.target_cost_per_usv, None)

    def test_target_cost_per_inq(self):
        self.assertEqual(self.period.target_cost_per_inq, None)

    def test_target_cost_per_tou(self):
        self.assertEqual(self.period.target_cost_per_tou, None)

    def test_target_cost_per_app(self):
        self.assertEqual(self.period.target_cost_per_app, None)

    def test_target_cost_per_exe(self):
        self.assertEqual(self.period.target_cost_per_exe, None)


class DefaultReportTestCase(TestCase):
    def setUp(self):
        project, _ = create_project()
        stage = LeaseStage.objects.get(short_name="performance")
        raw_period = Period.objects.create(
            project=project,
            lease_stage=stage,
            start=datetime.date(year=2018, month=12, day=19),
            end=datetime.date(year=2018, month=12, day=26),
            leased_units_start=0,
            occupiable_units_start=0,
            occupied_units_start=0,
        )
        raw_target_period = TargetPeriod.objects.create(
            project=project,
            start=datetime.date(year=2018, month=12, day=19),
            end=datetime.date(year=2018, month=12, day=26),
        )
        multiperiod = BareMultiPeriod.from_periods([raw_period, raw_target_period])
        period = multiperiod.get_cumulative_period()
        self.report = PerformanceReport(project, period)

    def test_report_jsonable(self):
        from django.core.serializers.json import DjangoJSONEncoder

        jsonable = self.report.to_jsonable()
        json_string = DjangoJSONEncoder().encode(jsonable)
        self.assertTrue(json_string)

    def test_report_jsonable_is_schema_valid(self):
        import json
        import jsonschema
        from django.core.serializers.json import DjangoJSONEncoder

        schema_location = os.path.join(
            os.path.dirname(os.path.abspath(__file__)),
            "./PerformanceReport.schema.json",
        )

        with open(schema_location, "rt") as schema_file:
            schema = json.load(schema_file)

        jsonable = self.report.to_jsonable()
        json_string = DjangoJSONEncoder().encode(jsonable)
        decoded_jsonable = json.loads(json_string)
        jsonschema.validate(instance=decoded_jsonable, schema=schema)


class LincolnTowerPeriodTestCase(TestCase):
    """Test an example Lincoln Tower period model with computed properties."""

    def setUp(self):
        address = Address.objects.create(
            street_address_1="2284 W. Commodore Way, Suite 200",
            city="Seattle",
            state="WA",
            zip_code=98199,
            country="US",
        )
        account = Account.objects.create(
            company_name="test", address=address, account_type=4
        )
        asset_manager = Business.objects.create(
            name="Test Asset Manager", is_asset_manager=True
        )
        property_manager = Business.objects.create(
            name="Test Property Manager", is_property_manager=True
        )
        property_owner = Business.objects.create(
            name="Test Property Owner", is_property_owner=True
        )
        fund = Fund.objects.create(account=account, name="Test Fund")
        property = Property.objects.create(
            name="test",
            total_units=220,
            average_monthly_rent=decimal.Decimal("7278"),
            lowest_monthly_rent=decimal.Decimal("7278"),
            geo_address=address,
        )
        self.project = Project.objects.create(
            name="test",
            baseline_start=datetime.date(year=2018, month=11, day=19),
            baseline_end=datetime.date(year=2018, month=12, day=26),
            account=account,
            asset_manager=asset_manager,
            property_manager=property_manager,
            property_owner=property_owner,
            fund=fund,
            property=property,
        )
        stage = LeaseStage.objects.get(short_name="performance")
        self.raw_period = Period.objects.create(
            project=self.project,
            lease_stage=stage,
            start=datetime.date(year=2018, month=12, day=19),
            end=datetime.date(year=2018, month=12, day=26),
            leased_units_start=104,
            usvs=4086,
            inquiries=51,
            tours=37,
            lease_applications=8,
            leases_executed=6,
            occupiable_units_start=218,
            occupied_units_start=218,
            leases_ended=3,
            lease_renewal_notices=0,
            acq_reputation_building=decimal.Decimal("28000"),
            acq_demand_creation=decimal.Decimal("21000"),
            acq_leasing_enablement=decimal.Decimal("11000"),
            acq_market_intelligence=decimal.Decimal("7000"),
        )

        self.raw_target_period = TargetPeriod.objects.create(
            project=self.project,
            start=datetime.date(year=2018, month=12, day=19),
            end=datetime.date(year=2018, month=12, day=26),
            target_leased_rate=decimal.Decimal("0.9"),
        )

        self.raw_multiperiod = BareMultiPeriod.from_periods(
            [self.raw_period, self.raw_target_period]
        )
        self.raw_cumulative = self.raw_multiperiod.get_cumulative_period()
        self.period = ComputedPeriod(self.raw_cumulative)

    def test_delta_leases(self):
        self.assertEqual(self.period.delta_leases, 3)

    def test_leased_units(self):
        self.assertEqual(self.period.leased_units, 107)

    def test_target_leased_units(self):
        self.assertEqual(self.period.target_leased_units, 196)

    def test_total_units(self):
        # total_units should be pulled from ref property
        self.assertEqual(self.period.total_units, 220)

    def test_leased_rate(self):
        self.assertEqual(round(self.period.leased_rate, 4), 0.4864)

    def test_occupancy_rate(self):
        self.assertEqual(round(self.period.occupancy_rate, 4), 0.9909)

    def test_usv_inq_perc(self):
        self.assertEqual(self.period.usv_inq_perc, 0.012481644640234948)

    def test_inq_tou_perc(self):
        self.assertEqual(self.period.inq_tou_perc, 0.7254901960784313)

    def test_tou_app_perc(self):
        self.assertEqual(self.period.tou_app_perc, 0.21621621621621623)

    def test_app_exe_perc(self):
        self.assertEqual(self.period.app_exe_perc, 0.75)

    def test_investment(self):
        self.assertEqual(self.period.investment, decimal.Decimal("67000"))

    def test_acq_investment(self):
        self.assertEqual(self.period.acq_investment, decimal.Decimal("67000"))

    def test_acq_investment_without_leasing(self):
        self.assertEqual(
            self.period.acq_investment_without_leasing, decimal.Decimal("56000")
        )

    def test_estimated_revenue_gain(self):
        self.assertEqual(self.period.estimated_revenue_gain, decimal.Decimal("262008"))

    def test_romi(self):
        self.assertEqual(self.period.romi, 4)

    def test_cost_per_usv(self):
        self.assertEqual(self.period.cost_per_usv, decimal.Decimal("13.71"))

    def test_cost_per_inq(self):
        self.assertEqual(self.period.cost_per_inq, decimal.Decimal("1098.04"))

    def test_cost_per_tou(self):
        self.assertEqual(self.period.cost_per_tou, decimal.Decimal("1810.81"))

    def test_cost_per_app(self):
        self.assertEqual(self.period.cost_per_app, decimal.Decimal("8375.00"))

    def test_cost_per_exe(self):
        self.assertEqual(self.period.cost_per_exe, decimal.Decimal("11166.67"))

    def test_report_jsonable(self):
        # CONSIDER moving this to a separate location
        report = PerformanceReport(self.project, self.raw_cumulative)
        self.assertTrue(report.to_jsonable())


from .signals import model_percent, get_ctd_top_kpis, sort_kpis, get_ctd_rest


class PerformanceEmailSignalTestCase(TestCase):
    """Test an example performace creation."""

    def setUp(self):
        pass

    def generate_mp_json(self, selectors, value, target):
        tl_values = {}
        values = tl_values
        tl_targets = {}
        targets = tl_targets
        for x in range(len(selectors)):
            if x >= len(selectors) - 1:
                values[selectors[x]] = value
                targets[selectors[x]] = target
            else:
                values[selectors[x]] = {}
                values = values[selectors[x]]
                targets[selectors[x]] = {}
                targets = targets[selectors[x]]
        tl_values["targets"] = tl_targets
        return tl_values

    def test_model_percent_equal(self):
        selectors = ["property", "leasing", "rate"]
        target = 0.80
        value = 0.80
        json_report = self.generate_mp_json(selectors, value, target)
        result = model_percent("lease_rate", json_report)
        self.assertEqual(result, 1.0)

    def test_model_percent_less(self):
        selectors = ["property", "leasing", "rate"]
        target = 0.80
        value = 0.40
        json_report = self.generate_mp_json(selectors, value, target)
        result = model_percent("lease_rate", json_report)
        self.assertEqual(result, 0.5)

    def test_model_percent_greater(self):
        selectors = ["property", "leasing", "rate"]
        target = 0.80
        value = 0.90
        json_report = self.generate_mp_json(selectors, value, target)
        result = model_percent("lease_rate", json_report)
        self.assertEqual(result, 1.125)

    def test_model_percent_move_outs_less(self):
        selectors = ["property", "occupancy", "move_outs"]
        target = 10
        value = 12
        json_report = self.generate_mp_json(selectors, value, target)
        result = model_percent("move_outs", json_report)
        self.assertEqual(result, 10 / 12)

    def test_model_percent_move_outs_greater(self):
        selectors = ["property", "occupancy", "move_outs"]
        target = 10
        value = 8
        json_report = self.generate_mp_json(selectors, value, target)
        result = model_percent("move_outs", json_report)
        self.assertEqual(result, 1.25)

    def test_get_ctd_top_kpis(self):
        ctd_model_percent = {"move_ins": 1.0, "move_outs": 0.6}
        ctd_sorted = sort_kpis(ctd_model_percent)
        result = get_ctd_top_kpis(ctd_model_percent, ctd_sorted)
        self.assertEqual(len(result), 1)
        self.assertEqual(result[0], "move_ins")

    def test_get_ctd_top_kpis_full_set(self):
        ctd_model_percent = {
            "move_ins": 1.0,
            "usv": 1.20,
            "inq": 1.40,
            "move_outs": 0.6,
        }
        ctd_sorted = sort_kpis(ctd_model_percent)
        result = get_ctd_top_kpis(ctd_model_percent, ctd_sorted)
        self.assertEqual(len(result), 3)
        self.assertEqual(result[0], "inq")
        self.assertEqual(result[1], "usv")
        self.assertEqual(result[2], "move_ins")

    def test_get_ctd_top_kpis_empty_set(self):
        ctd_model_percent = {
            "move_ins": 0.85,
            "usv": 0.20,
            "inq": 0.40,
            "move_outs": 0.6,
        }
        ctd_sorted = sort_kpis(ctd_model_percent)
        result = get_ctd_top_kpis(ctd_model_percent, ctd_sorted)
        self.assertEqual(len(result), 0)

    def test_get_ctd_rest(self):
        ctd_model_percent = {
            "move_ins": 0.95,
            "usv": 0.20,
            "inq": 0.40,
            "move_outs": 0.8,
        }
        ctd_sorted = sort_kpis(ctd_model_percent)
        risk, low = get_ctd_rest(ctd_model_percent, ctd_sorted)
        self.assertEqual(len(risk), 1)
        self.assertEqual(risk[0], "move_outs")
        self.assertEqual(len(low), 2)
        self.assertEqual(low[0], "usv")
        self.assertEqual(low[1], "inq")


class ExportTestCase(TestCase):
    def setUp(self):
        project, _ = create_project()
        stage = LeaseStage.objects.get(short_name="performance")
        raw_period = Period.objects.create(
            project=project,
            lease_stage=stage,
            start=datetime.date(year=2018, month=12, day=19),
            end=datetime.date(year=2018, month=12, day=26),
            leased_units_start=104,
            usvs=4086,
            inquiries=51,
            tours=37,
            lease_applications=8,
            leases_executed=6,
            occupiable_units_start=218,
            occupied_units_start=218,
            leases_ended=3,
            lease_renewal_notices=0,
            acq_reputation_building=decimal.Decimal("28000"),
            acq_demand_creation=decimal.Decimal("21000"),
            acq_leasing_enablement=decimal.Decimal("11000"),
            acq_market_intelligence=decimal.Decimal("7000.0"),
        )
        self.project = project
        self.period = raw_period

    def test_export_periods_to_csv(self):
        periods_ids = [self.period.id]
        response = export_periods_to_csv(periods_ids, self.project.public_id)
        csv_str = response.content.decode("utf-8")
        result = (
            "lease_stage,start,end,includes_remarkably_effect,"
            "leased_units_start,leases_ended,lease_applications,"
            "leases_executed,lease_cds,lease_renewal_notices,"
            "lease_renewals,lease_vacation_notices,"
            "occupiable_units_start,occupied_units_start,move_ins,"
            "move_outs,acq_reputation_building,acq_demand_creation,"
            "acq_leasing_enablement,acq_market_intelligence,"
            "ret_reputation_building,ret_demand_creation,"
            "ret_leasing_enablement,ret_market_intelligence,usvs,"
            "inquiries,tours"
            "\r\nImprove Performance,2018-12-19,2018-12-26,True,104,"
            "3,8,6,0,0,0,0,218,218,0,0,28000.00,21000.00,11000.00,"
            "7000.00,0.00,0.00,0.00,0.00,4086,51,37\r\n"
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(csv_str, result)

    @mock.patch("remark.projects.export.datetime")
    def test_export_periods_to_excel(self, mock_datetime):
        mock_datetime.date.today.return_value = datetime.datetime(2019, 10, 11, 0, 0)
        response = export_periods_to_excel(self.project.public_id)
        self.assertEqual(response.status_code, 200)

        response_wb = load_workbook(BytesIO(response.content))
        response_ws = response_wb["periods"]

        wb = load_workbook(filename="remark/projects/tests/periods.xlsx")
        ws_periods = wb["periods"]
        for row in range(1, response_ws.max_row + 1):
            for col in range(1, response_ws.max_column + 1):
                self.assertEqual(
                    ws_periods.cell(row=row, column=col).value,
                    response_ws.cell(row=row, column=col).value,
                    f"row: {row}, column: {col}",
                )


class OnboardingWorkflowTestCase(TestCase):
    def setUp(self):
        user = User.objects.create_user(
            email="admin@remarkably.io", password="adminpassword"
        )
        project, _ = create_project()
        admin_group = Group.objects.create(name="project 1 admin group")
        admin_group.user_set.add(user)
        project.admin_group = admin_group
        project.save()

        self.client.login(email="admin@remarkably.io", password="adminpassword")
        self.project = project
        self.user = user
        self.url = reverse("add_members")

    @mock.patch("remark.users.views.geocode", side_effect=mocked_geocode)
    @mock.patch("remark.projects.views.send_create_account_email.apply_async")
    @mock.patch(
        "remark.projects.views.send_invite_email.apply_async",
        side_effect=send_invite_email.apply,
    )
    @mock.patch("remark.email_app.invites.added_to_property.send_email")
    def test_invite_new_user(self, mock_send_email, *args):
        params = {
            "projects": [{"property_id": self.project.public_id}],
            "members": [
                {
                    "label": "new.user@gmail.com",
                    "value": "new.user@gmail.com",
                    "__isNew__": True,
                }
            ],
        }
        response = self.client.post(self.url, json.dumps(params), "json")
        self.assertEqual(response.status_code, 200)
        data = response.json()
        public_id = data["projects"][0]["members"][0]["user_id"]
        email = data["projects"][0]["members"][0]["email"]
        user = User.objects.get(public_id=public_id)
        self.assertEqual(email, user.email)
        self.assertFalse(user.password)
        mock_send_email.assert_called()

        self.client.logout()
        url = reverse("create_password", kwargs={"hash": public_id})
        params = {"password": "new_user_password"}
        response = self.client.post(url, json.dumps(params), "json")
        self.assertEqual(response.status_code, 200)
        data = response.json()
        redirect_url = reverse("complete_account")
        self.assertEqual(data["redirect_url"], redirect_url)
        user = User.objects.get(public_id=public_id)
        user.check_password(params["password"])
        self.assertTrue(user.check_password(params["password"]))

        params = {
            "first_name": "First Name",
            "last_name": "Last Name",
            "title": "Title",
            "company": "Company Name",
            "company_role": ["owner"],
            "office_address": "2284 W. Commodore Way, Suite 200",
            "office_name": "Office Name",
            "office_type": 3,
            "terms": True,
        }
        response = self.client.post(redirect_url, json.dumps(params), "json")
        self.assertEqual(response.status_code, 200)
        user = User.objects.get(public_id=public_id)
        self.assertTrue(user.person)
        self.assertTrue(user.person.office.address)

        project = Project.objects.get(public_id=self.project.public_id)
        project_users = project.view_group.user_set.all()
        self.assertEqual(project_users[0].public_id, user.public_id)

    @mock.patch("remark.projects.views.send_create_account_email.apply_async")
    @mock.patch(
        "remark.projects.views.send_invite_email.apply_async",
        side_effect=send_invite_email.apply,
    )
    @mock.patch("remark.email_app.invites.added_to_property.send_email")
    def test_invite_existing_user(self, mock_send_email, *args):
        user = User.objects.create_user(
            email="test@remarkably.io",
            password="testpassword",
            activated=datetime.datetime(2019, 10, 11, 0, 0),
        )
        params = {
            "projects": [{"property_id": self.project.public_id}],
            "members": [
                {"label": user.email, "value": user.public_id, "__isNew__": False}
            ],
        }
        response = self.client.post(self.url, json.dumps(params), "json")

        self.assertEqual(response.status_code, 200)
        mock_send_email.assert_called()

        project_users = self.project.view_group.user_set.all()
        self.assertEqual(project_users[0].public_id, user.public_id)


class GetTemplateVarsTestCase(TestCase):
    def setUp(self):
        project1, _ = create_project("project 1")
        project2, _ = create_project("project 2")
        user = User.objects.create_user(
            email="test@remarkably.io",
            password="testpassword",
            activated=datetime.datetime(2019, 10, 11, 0, 0),
        )
        new_user = User.objects.create_user(
            email="test_new@remarkably.io", password="testpassword"
        )
        self.user = user
        self.new_user = new_user
        self.project = project1
        self.projects = [project1, project2]

    def test_for_new_user(self):
        template_vars = get_template_vars("admin", self.new_user, [self.project], 5)
        expected = {
            "email_title": "Added to New Property",
            "email_preview": "Added to New Property",
            "inviter_name": "admin",
            "is_portfolio": False,
            "is_new_account": True,
            "property_name": "project 1",
            "properties": [
                {
                    "image_url": "https://s3.amazonaws.com/production-storage.remarkably.io/email_assets/blank_property_square.png",
                    "title": "project 1",
                    "address": "Seattle, WA",
                    "view_link": f"{BASE_URL}/projects/{self.project.public_id}/market/",
                }
            ],
            "more_count": None,
            "main_button_link": f"{BASE_URL}/users/create-password/{self.new_user.public_id}",
            "main_button_label": "Create Account",
        }
        self.assertEqual(expected, template_vars)

    def test_for_new_user_many_projects(self):
        template_vars = get_template_vars("admin", self.new_user, self.projects, 5)
        expected = {
            "email_title": "Added to New Property",
            "email_preview": "Added to New Property",
            "inviter_name": "admin",
            "is_portfolio": False,
            "is_new_account": True,
            "property_name": "",
            "properties": [
                {
                    "image_url": "https://s3.amazonaws.com/production-storage.remarkably.io/email_assets/blank_property_square.png",
                    "title": "project 1",
                    "address": "Seattle, WA",
                    "view_link": f"{BASE_URL}/projects/{self.projects[0].public_id}/market/",
                },
                {
                    "image_url": "https://s3.amazonaws.com/production-storage.remarkably.io/email_assets/blank_property_square.png",
                    "title": "project 2",
                    "address": "Seattle, WA",
                    "view_link": f"{BASE_URL}/projects/{self.projects[1].public_id}/market/",
                },
            ],
            "more_count": None,
            "main_button_link": f"{BASE_URL}/users/create-password/{self.new_user.public_id}",
            "main_button_label": "Create Account",
        }
        self.assertEqual(expected, template_vars)

    def test_for_existing_user(self):
        template_vars = get_template_vars("admin", self.user, [self.project], 5)
        expected = {
            "email_title": "Added to New Property",
            "email_preview": "Added to New Property",
            "inviter_name": "admin",
            "is_portfolio": False,
            "is_new_account": False,
            "property_name": "project 1",
            "properties": [
                {
                    "image_url": "https://s3.amazonaws.com/production-storage.remarkably.io/email_assets/blank_property_square.png",
                    "title": "project 1",
                    "address": "Seattle, WA",
                    "view_link": f"{BASE_URL}/projects/{self.project.public_id}/market/",
                }
            ],
            "more_count": None,
            "main_button_link": f"{BASE_URL}/projects/{self.project.public_id}/market/",
            "main_button_label": "View Property",
        }
        self.assertEqual(expected, template_vars)

    def test_for_existing_user_many_projects(self):
        template_vars = get_template_vars("admin", self.user, self.projects, 5)
        expected = {
            "email_title": "Added to New Property",
            "email_preview": "Added to New Property",
            "inviter_name": "admin",
            "is_portfolio": False,
            "is_new_account": False,
            "property_name": "",
            "properties": [
                {
                    "image_url": "https://s3.amazonaws.com/production-storage.remarkably.io/email_assets/blank_property_square.png",
                    "title": "project 1",
                    "address": "Seattle, WA",
                    "view_link": f"{BASE_URL}/projects/{self.projects[0].public_id}/market/",
                },
                {
                    "image_url": "https://s3.amazonaws.com/production-storage.remarkably.io/email_assets/blank_property_square.png",
                    "title": "project 2",
                    "address": "Seattle, WA",
                    "view_link": f"{BASE_URL}/projects/{self.projects[1].public_id}/market/",
                },
            ],
            "more_count": None,
            "main_button_link": f"{BASE_URL}/dashboard",
            "main_button_label": "View All Properties",
        }
        self.maxDiff = None
        self.assertEqual(expected, template_vars)


class AutocompleteMemberTestCase(TestCase):
    @staticmethod
    def add_users_to_group(users, group):
        for user in users:
            group.user_set.add(user)

    @staticmethod
    def generate_users(start=0, end=5):
        users = []
        for i in range(start, end):
            email = f"user{i}@test.local"
            user = User.objects.create_user(email=email, password="password")
            users.append(user)
        return users

    def setUp(self):
        user = User.objects.create_user(
            email="admin@remarkably.io", password="adminpassword"
        )
        project, group = create_project()
        users = AutocompleteMemberTestCase.generate_users(0, 5)
        AutocompleteMemberTestCase.add_users_to_group(users, group)
        self.client.login(email="admin@remarkably.io", password="adminpassword")
        self.user = user
        self.group = group
        self.users = users
        self.url = reverse("members")

    def test_without_projects(self):
        params = {"value": ""}
        response = self.client.post(self.url, json.dumps(params), "json")
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data["members"], [])

    def test_with_project(self):
        self.group.user_set.add(self.user)
        params = {"value": ""}
        response = self.client.post(self.url, json.dumps(params), "json")
        self.assertEqual(response.status_code, 200)
        data = response.json()
        users = [u.get_icon_dict() for u in self.users]
        self.assertCountEqual(data["members"], users)

    def test_many_projects(self):
        self.group.user_set.add(self.user)
        _, group = create_project("project 2")
        users = AutocompleteMemberTestCase.generate_users(5, 10)
        AutocompleteMemberTestCase.add_users_to_group(users, group)
        params = {"value": ""}
        response = self.client.post(self.url, json.dumps(params), "json")
        self.assertEqual(response.status_code, 200)
        data = response.json()
        users = [u.get_icon_dict() for u in self.users]
        self.assertCountEqual(data["members"], users)

    def test_many_projects_with_user(self):
        AutocompleteMemberTestCase.add_users_to_group([self.user], self.group)
        _, group = create_project("project 2")
        AutocompleteMemberTestCase.add_users_to_group(self.users, group)
        AutocompleteMemberTestCase.add_users_to_group([self.user], group)
        params = {"value": ""}
        response = self.client.post(self.url, json.dumps(params), "json")
        self.assertEqual(response.status_code, 200)
        data = response.json()
        users = [u.get_icon_dict() for u in self.users]
        self.assertCountEqual(data["members"], users)

    def test_select_staff(self):
        self.group.user_set.add(self.user)
        staff_user = self.users[0]
        staff_user.is_staff = True
        staff_user.save()

        params = {"value": staff_user.email}
        response = self.client.post(self.url, json.dumps(params), "json")
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertFalse(data["members"])

        params = {"value": ""}
        response = self.client.post(self.url, json.dumps(params), "json")
        self.assertEqual(response.status_code, 200)
        data = response.json()
        users = [u.get_icon_dict() for u in self.users[1:]]
        self.assertCountEqual(data["members"], users)


class AddMembersTestCase(TestCase):
    def setUp(self):
        self.url = reverse("add_members")
        User.objects.create_user(
            email="admin@remarkably.io", password="adminpassword", is_superuser=True
        )
        self.client.login(email="admin@remarkably.io", password="adminpassword")

    def get_params_for_new_users(self, projects, emails=None):
        if emails is None:
            emails = ["new_user@remarkably.io"]
        return {
            "projects": [{"property_id": p.public_id} for p in projects],
            "members": [{"__isNew__": True, "value": email} for email in emails],
            "role": "member",
        }

    def get_params_for_existing_users(self, projects, emails=None):
        if emails is None:
            emails = ["invited_user@remarkably.io"]
        users = []
        for email in emails:
            user = User.objects.create(
                email=email, activated=datetime.datetime.now(tz=timezone.utc)
            )
            user.save()
            users.append(user)
        return {
            "projects": [{"property_id": p.public_id} for p in projects],
            "members": [{"__isNew__": False, "value": u.public_id} for u in users],
            "role": "member",
        }

    def get_params_for_mix_users(self, projects):
        user = User.objects.create(
            email="invited_user@remarkably.io",
            activated=datetime.datetime.now(tz=timezone.utc),
        )
        return {
            "projects": [{"property_id": p.public_id} for p in projects],
            "members": [
                {"__isNew__": False, "value": user.public_id},
                {"__isNew__": True, "value": "new_user@remarkably.io"},
            ],
            "role": "member",
        }

    @mock.patch("remark.projects.views.send_create_account_email.apply_async")
    @mock.patch("remark.projects.views.send_invite_email.apply_async")
    def test_new_user_to_one_property_without_users(
        self, mock_send_invite_email, mock_send_create_account_email
    ):
        project, _ = create_project()
        params = self.get_params_for_new_users([project])
        response = self.client.post(self.url, json.dumps(params), "json")
        self.assertEqual(response.status_code, 200)
        mock_send_invite_email.assert_not_called()
        mock_send_create_account_email.assert_called_once()

    @mock.patch("remark.projects.views.send_create_account_email.apply_async")
    @mock.patch("remark.projects.views.send_invite_email.apply_async")
    def test_new_user_to_one_property_with_user(
        self, mock_send_invite_email, mock_send_create_account_email
    ):
        project, group = create_project()
        add_user_to_group(group)
        params = self.get_params_for_new_users([project])
        response = self.client.post(self.url, json.dumps(params), "json")
        self.assertEqual(response.status_code, 200)
        mock_send_invite_email.assert_called_once()
        mock_send_create_account_email.assert_not_called()

    @mock.patch("remark.projects.views.send_create_account_email.apply_async")
    @mock.patch("remark.projects.views.send_invite_email.apply_async")
    def test_new_user_to_many_properties_without_users(
        self, mock_send_invite_email, mock_send_create_account_email
    ):
        project1, _ = create_project("project 1")
        project2, _ = create_project("project 2")
        params = self.get_params_for_new_users([project1, project2])
        response = self.client.post(self.url, json.dumps(params), "json")
        self.assertEqual(response.status_code, 200)
        mock_send_invite_email.assert_not_called()
        mock_send_create_account_email.assert_called_once()

    @mock.patch("remark.projects.views.send_create_account_email.apply_async")
    @mock.patch("remark.projects.views.send_invite_email.apply_async")
    def test_new_user_to_many_properties_with_users(
        self, mock_send_invite_email, mock_send_create_account_email
    ):
        project1, group1 = create_project("project 1")
        add_user_to_group(group1)
        project2, group2 = create_project("project 2")
        add_user_to_group(group2)
        params = self.get_params_for_new_users([project1, project2])
        response = self.client.post(self.url, json.dumps(params), "json")
        self.assertEqual(response.status_code, 200)
        mock_send_invite_email.assert_called_once()
        mock_send_create_account_email.assert_not_called()

    @mock.patch("remark.projects.views.send_create_account_email.apply_async")
    @mock.patch("remark.projects.views.send_invite_email.apply_async")
    def test_new_user_to_many_properties(
        self, mock_send_invite_email, mock_send_create_account_email
    ):
        project1, _ = create_project("project 1")
        project2, group2 = create_project("project 2")
        add_user_to_group(group2)
        params = self.get_params_for_new_users([project1, project2])
        response = self.client.post(self.url, json.dumps(params), "json")
        self.assertEqual(response.status_code, 200)
        mock_send_invite_email.assert_not_called()
        mock_send_create_account_email.assert_called_once()

    @mock.patch("remark.projects.views.send_create_account_email.apply_async")
    @mock.patch("remark.projects.views.send_invite_email.apply_async")
    def test_user_to_one_property_without_users(
        self, mock_send_invite_email, mock_send_create_account_email
    ):
        project, _ = create_project()
        params = self.get_params_for_existing_users([project])
        response = self.client.post(self.url, json.dumps(params), "json")
        self.assertEqual(response.status_code, 200)
        mock_send_invite_email.assert_called_once()
        mock_send_create_account_email.assert_not_called()

    @mock.patch("remark.projects.views.send_create_account_email.apply_async")
    @mock.patch("remark.projects.views.send_invite_email.apply_async")
    def test_user_to_one_property_with_user(
        self, mock_send_invite_email, mock_send_create_account_email
    ):
        project, group = create_project()
        add_user_to_group(group)
        params = self.get_params_for_existing_users([project])
        response = self.client.post(self.url, json.dumps(params), "json")
        self.assertEqual(response.status_code, 200)
        mock_send_invite_email.assert_called_once()
        mock_send_create_account_email.assert_not_called()

    @mock.patch("remark.projects.views.send_create_account_email.apply_async")
    @mock.patch("remark.projects.views.send_invite_email.apply_async")
    def test_user_to_many_properties_without_users(
        self, mock_send_invite_email, mock_send_create_account_email
    ):
        project1, _ = create_project("project 1")
        project2, _ = create_project("project 2")
        params = self.get_params_for_existing_users([project1, project2])
        response = self.client.post(self.url, json.dumps(params), "json")
        self.assertEqual(response.status_code, 200)
        mock_send_invite_email.assert_called_once()
        mock_send_create_account_email.assert_not_called()

    @mock.patch("remark.projects.views.send_create_account_email.apply_async")
    @mock.patch("remark.projects.views.send_invite_email.apply_async")
    def test_user_to_many_properties_with_users(
        self, mock_send_invite_email, mock_send_create_account_email
    ):
        project1, group1 = create_project("project 1")
        add_user_to_group(group1)
        project2, group2 = create_project("project 2")
        add_user_to_group(group2)
        params = self.get_params_for_existing_users([project1, project2])
        response = self.client.post(self.url, json.dumps(params), "json")
        self.assertEqual(response.status_code, 200)
        mock_send_invite_email.assert_called_once()
        mock_send_create_account_email.assert_not_called()

    @mock.patch("remark.projects.views.send_create_account_email.apply_async")
    @mock.patch("remark.projects.views.send_invite_email.apply_async")
    def test_user_to_many_properties(
        self, mock_send_invite_email, mock_send_create_account_email
    ):
        project1, _ = create_project("project 1")
        project2, group2 = create_project("project 2")
        add_user_to_group(group2)
        params = self.get_params_for_existing_users([project1, project2])
        response = self.client.post(self.url, json.dumps(params), "json")
        self.assertEqual(response.status_code, 200)
        mock_send_invite_email.assert_called_once()
        mock_send_create_account_email.assert_not_called()

    @mock.patch("remark.projects.views.send_create_account_email.apply_async")
    @mock.patch("remark.projects.views.send_invite_email.apply_async")
    def test_new_users_to_one_property_without_users(
        self, mock_send_invite_email, mock_send_create_account_email
    ):
        project, _ = create_project()
        params = self.get_params_for_new_users(
            [project], ["new_user1@remarkably.io", "new_user2@remarkably.io"]
        )
        response = self.client.post(self.url, json.dumps(params), "json")
        self.assertEqual(response.status_code, 200)
        mock_send_invite_email.assert_not_called()
        mock_send_create_account_email.assert_called()
        self.assertEqual(mock_send_create_account_email.call_count, 2)

    @mock.patch("remark.projects.views.send_create_account_email.apply_async")
    @mock.patch("remark.projects.views.send_invite_email.apply_async")
    def test_new_users_to_one_property_with_user(
        self, mock_send_invite_email, mock_send_create_account_email
    ):
        project, group = create_project()
        add_user_to_group(group)
        params = self.get_params_for_new_users(
            [project], ["new_user1@remarkably.io", "new_user2@remarkably.io"]
        )
        response = self.client.post(self.url, json.dumps(params), "json")
        self.assertEqual(response.status_code, 200)
        mock_send_invite_email.assert_called()
        self.assertEqual(mock_send_invite_email.call_count, 2)
        mock_send_create_account_email.assert_not_called()

    @mock.patch("remark.projects.views.send_create_account_email.apply_async")
    @mock.patch("remark.projects.views.send_invite_email.apply_async")
    def test_new_users_to_many_properties_without_users(
        self, mock_send_invite_email, mock_send_create_account_email
    ):
        project1, _ = create_project("project 1")
        project2, _ = create_project("project 2")
        params = self.get_params_for_new_users(
            [project1, project2], ["new_user1@remarkably.io", "new_user2@remarkably.io"]
        )
        response = self.client.post(self.url, json.dumps(params), "json")
        self.assertEqual(response.status_code, 200)
        mock_send_invite_email.assert_not_called()
        mock_send_create_account_email.assert_called()
        self.assertEqual(mock_send_create_account_email.call_count, 2)

    @mock.patch("remark.projects.views.send_create_account_email.apply_async")
    @mock.patch("remark.projects.views.send_invite_email.apply_async")
    def test_new_users_to_many_properties_with_users(
        self, mock_send_invite_email, mock_send_create_account_email
    ):
        project1, group1 = create_project("project 1")
        add_user_to_group(group1)
        project2, group2 = create_project("project 2")
        add_user_to_group(group2)
        params = self.get_params_for_new_users(
            [project1, project2], ["new_user1@remarkably.io", "new_user2@remarkably.io"]
        )
        response = self.client.post(self.url, json.dumps(params), "json")
        self.assertEqual(response.status_code, 200)
        mock_send_invite_email.assert_called()
        self.assertEqual(mock_send_invite_email.call_count, 2)
        mock_send_create_account_email.assert_not_called()

    @mock.patch("remark.projects.views.send_create_account_email.apply_async")
    @mock.patch("remark.projects.views.send_invite_email.apply_async")
    def test_new_users_to_many_properties(
        self, mock_send_invite_email, mock_send_create_account_email
    ):
        project1, _ = create_project("project 1")
        project2, group2 = create_project("project 2")
        add_user_to_group(group2)
        params = self.get_params_for_new_users(
            [project1, project2], ["new_user1@remarkably.io", "new_user2@remarkably.io"]
        )
        response = self.client.post(self.url, json.dumps(params), "json")
        self.assertEqual(response.status_code, 200)
        mock_send_invite_email.assert_not_called()
        mock_send_create_account_email.assert_called()
        self.assertEqual(mock_send_create_account_email.call_count, 2)

    @mock.patch("remark.projects.views.send_create_account_email.apply_async")
    @mock.patch("remark.projects.views.send_invite_email.apply_async")
    def test_existing_users_to_one_property_without_users(
        self, mock_send_invite_email, mock_send_create_account_email
    ):
        project, _ = create_project()
        params = self.get_params_for_existing_users(
            [project], ["invited_user1@remarkably.io", "invited_user2@remarkably.io"]
        )
        response = self.client.post(self.url, json.dumps(params), "json")
        self.assertEqual(response.status_code, 200)
        mock_send_invite_email.assert_called()
        self.assertEqual(mock_send_invite_email.call_count, 2)
        mock_send_create_account_email.assert_not_called()

    @mock.patch("remark.projects.views.send_create_account_email.apply_async")
    @mock.patch("remark.projects.views.send_invite_email.apply_async")
    def test_existing_users_to_one_property_with_user(
        self, mock_send_invite_email, mock_send_create_account_email
    ):
        project, group = create_project()
        add_user_to_group(group)
        params = self.get_params_for_existing_users(
            [project], ["invited_user1@remarkably.io", "invited_user2@remarkably.io"]
        )
        response = self.client.post(self.url, json.dumps(params), "json")
        self.assertEqual(response.status_code, 200)
        mock_send_invite_email.assert_called()
        self.assertEqual(mock_send_invite_email.call_count, 2)
        mock_send_create_account_email.assert_not_called()

    @mock.patch("remark.projects.views.send_create_account_email.apply_async")
    @mock.patch("remark.projects.views.send_invite_email.apply_async")
    def test_existing_users_to_many_properties_without_users(
        self, mock_send_invite_email, mock_send_create_account_email
    ):
        project1, _ = create_project("project 1")
        project2, _ = create_project("project 2")
        params = self.get_params_for_existing_users(
            [project1, project2],
            ["invited_user1@remarkably.io", "invited_user2@remarkably.io"],
        )
        response = self.client.post(self.url, json.dumps(params), "json")
        self.assertEqual(response.status_code, 200)
        mock_send_invite_email.assert_called()
        self.assertEqual(mock_send_invite_email.call_count, 2)
        mock_send_create_account_email.assert_not_called()

    @mock.patch("remark.projects.views.send_create_account_email.apply_async")
    @mock.patch("remark.projects.views.send_invite_email.apply_async")
    def test_existing_users_to_many_properties_with_users(
        self, mock_send_invite_email, mock_send_create_account_email
    ):
        project1, group1 = create_project("project 1")
        add_user_to_group(group1)
        project2, group2 = create_project("project 2")
        add_user_to_group(group2)
        params = self.get_params_for_existing_users(
            [project1, project2],
            ["invited_user1@remarkably.io", "invited_user2@remarkably.io"],
        )
        response = self.client.post(self.url, json.dumps(params), "json")
        self.assertEqual(response.status_code, 200)
        mock_send_invite_email.assert_called()
        self.assertEqual(mock_send_invite_email.call_count, 2)
        mock_send_create_account_email.assert_not_called()

    @mock.patch("remark.projects.views.send_create_account_email.apply_async")
    @mock.patch("remark.projects.views.send_invite_email.apply_async")
    def test_existing_users_to_many_properties(
        self, mock_send_invite_email, mock_send_create_account_email
    ):
        project1, _ = create_project("project 1")
        project2, group2 = create_project("project 2")
        add_user_to_group(group2)
        params = self.get_params_for_existing_users(
            [project1, project2],
            ["invited_user1@remarkably.io", "invited_user2@remarkably.io"],
        )
        response = self.client.post(self.url, json.dumps(params), "json")
        self.assertEqual(response.status_code, 200)
        mock_send_invite_email.assert_called()
        self.assertEqual(mock_send_invite_email.call_count, 2)
        mock_send_create_account_email.assert_not_called()

    @mock.patch("remark.projects.views.send_create_account_email.apply_async")
    @mock.patch("remark.projects.views.send_invite_email.apply_async")
    def test_users_to_one_property_without_users(
        self, mock_send_invite_email, mock_send_create_account_email
    ):
        project, _ = create_project()
        params = self.get_params_for_mix_users([project])
        response = self.client.post(self.url, json.dumps(params), "json")
        self.assertEqual(response.status_code, 200)
        mock_send_invite_email.assert_called_once()
        mock_send_create_account_email.assert_called_once()

    @mock.patch("remark.projects.views.send_create_account_email.apply_async")
    @mock.patch("remark.projects.views.send_invite_email.apply_async")
    def test_users_to_one_property_with_user(
        self, mock_send_invite_email, mock_send_create_account_email
    ):
        project, group = create_project()
        add_user_to_group(group)
        params = self.get_params_for_mix_users([project])
        response = self.client.post(self.url, json.dumps(params), "json")
        self.assertEqual(response.status_code, 200)
        mock_send_invite_email.assert_called()
        self.assertEqual(mock_send_invite_email.call_count, 2)
        mock_send_create_account_email.assert_not_called()

    @mock.patch("remark.projects.views.send_create_account_email.apply_async")
    @mock.patch("remark.projects.views.send_invite_email.apply_async")
    def test_users_to_many_properties_without_users(
        self, mock_send_invite_email, mock_send_create_account_email
    ):
        project1, _ = create_project("project 1")
        project2, _ = create_project("project 2")
        params = self.get_params_for_mix_users([project1, project2])
        response = self.client.post(self.url, json.dumps(params), "json")
        self.assertEqual(response.status_code, 200)
        mock_send_invite_email.assert_called_once()
        mock_send_create_account_email.assert_called_once()

    @mock.patch("remark.projects.views.send_create_account_email.apply_async")
    @mock.patch("remark.projects.views.send_invite_email.apply_async")
    def test_users_to_many_properties_with_users(
        self, mock_send_invite_email, mock_send_create_account_email
    ):
        project1, group1 = create_project("project 1")
        add_user_to_group(group1)
        project2, group2 = create_project("project 2")
        add_user_to_group(group2)
        params = self.get_params_for_mix_users([project1, project2])
        response = self.client.post(self.url, json.dumps(params), "json")
        self.assertEqual(response.status_code, 200)
        mock_send_invite_email.assert_called()
        self.assertEqual(mock_send_invite_email.call_count, 2)
        mock_send_create_account_email.assert_not_called()

    @mock.patch("remark.projects.views.send_create_account_email.apply_async")
    @mock.patch("remark.projects.views.send_invite_email.apply_async")
    def test_users_to_many_properties(
        self, mock_send_invite_email, mock_send_create_account_email
    ):
        project1, _ = create_project("project 1")
        project2, group2 = create_project("project 2")
        add_user_to_group(group2)
        params = self.get_params_for_mix_users([project1, project2])
        response = self.client.post(self.url, json.dumps(params), "json")
        self.assertEqual(response.status_code, 200)
        mock_send_invite_email.assert_called_once()
        mock_send_create_account_email.assert_called_once()


class PropertyStyleTestCase(TestCase):
    def setUp(self):
        address = Address.objects.create(
            street_address_1="2284 W. Commodore Way, Suite 200",
            city="Seattle",
            state="WA",
            zip_code=98199,
            country="US",
        )
        self.property = Property.objects.create(
            name="Test Property",
            average_monthly_rent=decimal.Decimal("0"),
            lowest_monthly_rent=decimal.Decimal("0"),
            geo_address=address,
        )

    def test_empty_style(self):
        self.assertIsNone(self.property.property_style)

    def test_is_low_rise(self):
        build = Building(
            property=self.property,
            building_identifier="building identifier",
            number_of_floors=4,
            has_elevator=True,
            number_of_units=10,
        )
        build.save()

        self.assertEqual("Low-Rise", self.property.property_style)

    def test_is_walk_up(self):
        build = Building(
            property=self.property,
            building_identifier="building identifier",
            number_of_floors=4,
            has_elevator=False,
            number_of_units=10,
        )
        build.save()

        self.assertEqual("Walk-Up", self.property.property_style)

    def test_is_mid_rise(self):
        build = Building(
            property=self.property,
            building_identifier="building identifier",
            number_of_floors=9,
            has_elevator=False,
            number_of_units=10,
        )
        build.save()

        self.assertEqual("Mid-Rise", self.property.property_style)

    def test_is_hi_rise(self):
        build = Building(
            property=self.property,
            building_identifier="building identifier",
            number_of_floors=10,
            has_elevator=False,
            number_of_units=10,
        )
        build.save()

        self.assertEqual("Hi-Rise", self.property.property_style)

    def test_is_tower_block(self):
        build_1 = Building(
            property=self.property,
            building_identifier="building identifier 1",
            number_of_floors=10,
            has_elevator=False,
            number_of_units=10,
        )
        build_1.save()

        build_2 = Building(
            property=self.property,
            building_identifier="building identifier 2",
            number_of_floors=1,
            has_elevator=True,
            number_of_units=10,
        )
        build_2.save()

        self.assertEqual("Tower-Block", self.property.property_style)

    def test_is_garden(self):
        build_1 = Building(
            property=self.property,
            building_identifier="building identifier 1",
            number_of_floors=9,
            has_elevator=True,
            number_of_units=10,
        )
        build_1.save()

        build_2 = Building(
            property=self.property,
            building_identifier="building identifier 2",
            number_of_floors=1,
            has_elevator=False,
            number_of_units=10,
        )
        build_2.save()

        self.assertEqual("Garden", self.property.property_style)
