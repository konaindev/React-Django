import csv
import os
import datetime
from django.http import HttpResponse
from openpyxl import load_workbook
from openpyxl.writer.excel import save_virtual_workbook

from remark.lib.logging import getLogger

from .models import Period, Project

logger = getLogger(__name__)

EXPORT_CSV_FIELDS = [
    "lease_stage",
    "start",
    "end",
    "includes_remarkably_effect",
    "leased_units_start",
    "leases_ended",
    "lease_applications",
    "leases_executed",
    "lease_cds",
    "lease_renewal_notices",
    "lease_renewals",
    "lease_vacation_notices",
    "occupiable_units_start",
    "occupied_units_start",
    "move_ins",
    "move_outs",
    "acq_reputation_building",
    "acq_demand_creation",
    "acq_leasing_enablement",
    "acq_market_intelligence",
    "ret_reputation_building",
    "ret_demand_creation",
    "ret_leasing_enablement",
    "ret_market_intelligence",
    "usvs",
    "inquiries",
    "tours",
]


def export_periods_to_csv(periods_ids, project_id):
    periods = Period.objects.filter(project_id=project_id, id__in=periods_ids)
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="periods.csv"'
    writer = csv.writer(response)
    writer.writerow(EXPORT_CSV_FIELDS)
    for p in periods:
        row = []
        for f in EXPORT_CSV_FIELDS:
            row.append(getattr(p, f))
        writer.writerow(row)
    return response


def export_periods_to_excel(project_id):
    project = Project.objects.get(public_id=project_id)

    def _baseline_formatter(end_date):
        if end_date <= project.get_baseline_end():
            return "baseline"
        else:
            return ""

    TEMPLATE_FILE_NAME = os.path.join(
        os.path.abspath(os.path.dirname(__file__)),
        "spreadsheets/templates/baseline-perf.xlsx",
    )

    wb = load_workbook(TEMPLATE_FILE_NAME)
    ws_periods = wb.get_sheet_by_name("periods")

    fields = [
        {"name": "start"},
        {"name": "end", "formatter": _baseline_formatter},
        {"name": "lease_stage", "formatter": str},
        {"name": "leased_units_start"},
        {"name": "lease_applications"},
        {"name": "leases_executed"},
        {"name": "leases_ended"},
        {"name": "lease_cds"},
        {"name": "lease_renewals"},
        {"name": "lease_renewal_notices"},
        {"name": "lease_vacation_notices"},
        {"name": "occupied_units_start"},
        {"name": "occupiable_units_start"},
        {"name": "move_ins"},
        {"name": "move_outs"},
        {"name": "usvs"},
        {"name": "inquiries"},
        {"name": "tours"},
        {"name": "acq_reputation_building"},
        {"name": "acq_demand_creation"},
        {"name": "acq_leasing_enablement"},
        {"name": "acq_market_intelligence"},
        {"name": "ret_reputation_building"},
        {"name": "ret_demand_creation"},
        {"name": "ret_leasing_enablement"},
        {"name": "ret_market_intelligence"},
    ]

    periods = Period.objects.filter(project_id=project_id)
    last_row = 0
    for row, period in enumerate(periods, start=3):
        for col, field in enumerate(fields, start=1):
            value = getattr(period, field["name"])
            if "formatter" in field:
                value = field["formatter"](value)
            try:
                ws_periods.cell(column=col, row=row, value=value)
            except ValueError:
                logger.error(
                    f"Cannot export: Project ID: {project_id} | column: {col} | row: {row} | value: {value}"
                )
        last_row = row
    ws_periods.cell(column=1, row=last_row+1, value=datetime.date.today())

    response = HttpResponse(
        content=save_virtual_workbook(wb),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    wb.close()
    response["Content-Disposition"] = 'attachment; filename="periods.xlsx"'
    return response
