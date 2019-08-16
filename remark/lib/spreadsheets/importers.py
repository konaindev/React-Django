import openpyxl

from .errors import ExcelError, ExcelValidationError
from .getset import get_cell
from .rowcol import col_range, row_range, location_range, location_range_rect
from .parse import parse_location_or_default
from .schema import DataType, is_schema_cell


class ExcelImporter:
    """
    Provides bare-bones tools for importing *any* of our data entry spreadsheets.
    """

    def __init__(self, f):
        """
        Create an importer with a fileobj or with the path to an extant file.
        """
        # Do *not* use read_only=True because this forces openpyxl into a mode
        # where *every* sheet/cell access results in re-parsing the XML. That's...
        # insane, and a sign that openpyxl is problematic. -Dave
        #
        # Example timing for one of our spreadsheets with read_only=True:
        #     imported full spreadsheet in 6.220s
        #
        # Example timing for that same spreadsheet without:
        #     imported full spreadsheet in 0.410s
        #
        # Insane!
        self.workbook = openpyxl.load_workbook(f, data_only=True)
        self.cleaned_data = {}
        self.errors = []

    def check_data_type(self, cell, data_type, label):
        """Raise an exception if cell's data type doesn't match provided data_type."""
        # If we have no expectations, we're always happy
        if callable(data_type):
            valid = data_type(cell)
        elif data_type == DataType.DATETIME:
            valid = cell.is_date
        else:
            valid = data_type == cell.data_type
        if not valid:
            raise ExcelValidationError(
                f"For {label}: found data type '{cell.data_type}' but expected '{data_type}'",
                where=cell,
            )

    def check_convert(self, cell, converter=None, label=None):
        """
        Attempt to convert a cell's raw value to a python value.

        Raise an exception if the conversion fails. Otherwise, return
        the converted value.
        """
        # CONSIDER this is a pretty awkward API, I think?
        value = cell.value
        if converter is not None:
            try:
                value = converter(value)
            except Exception as e:
                raise ExcelValidationError(
                    f"For {label}: Could not apply value converter '{converter}': '{e}'", where=cell
                )
        return value

    def cell(self, schema_cell=None, location=None, sheet=None, col=None, row=None):
        """
        Return an openpyxl Cell instance from the spreadsheet.

        The location of the cell is determined first and foremost by the `location`
        provided here, and secondarily by the schema_cell's `locator` (if any).

        Typically, you'll *declare* how to find a cell by setting the schema_cell
        to a SchemaCell tuple.

        That said, if you just want to get a Cell from the spreadsheet directly
        and without any fussy schema mechanisms you can always call with
        (for example): self.cell(location="output!A3")
        """
        sheet, col, row = parse_location_or_default(location, sheet, col, row)
        if (schema_cell is not None) and callable(schema_cell.locator):
            sheet, col, row = schema_cell.locator(self.workbook, sheet, col, row)
        return get_cell(self.workbook, sheet, col, row)

    def value(self, schema_cell=None, location=None, sheet=None, col=None, row=None):
        """
        Return the value of a cell in the spreadsheet.

        The location of the cell is determined first and foremost by the `location`
        provided here, and secondarily by the schema_cell's `locator` (if any).

        Typically, you'll *declare* how to find a cell, and validate its value,
        by setting the schema_cell to a SchemaCell tuple.

        That said, if you just want to get a value from the spreadsheet directly
        and without any fussy schema mechanisms you can always call with
        (for example): self.value(location="output!A3")
        """
        cell = self.cell(schema_cell, location, sheet, col, row)
        if schema_cell is not None:
            self.check_data_type(cell, schema_cell.data_type, schema_cell.label)
            value = self.check_convert(cell, schema_cell.converter, schema_cell.label)
        else:
            value = cell.value
        return value

    def check_value(
        self,
        schema_cell=None,
        location=None,
        sheet=None,
        col=None,
        row=None,
        expected=None,
    ):
        """
        Raise an exception if the wrong value is found.

        `expected` can be an explicit value, or a callable; if a callable
        is provided, it is called with the converted value and must return
        True if the expectation is met.
        """
        cell = self.cell(schema_cell, location, sheet, col, row)
        self.check_data_type(cell, schema_cell.data_type, schema_cell.label)
        value = self.check_convert(cell, schema_cell.converter, schema_cell.label)
        value_matches = expected(value) if callable(expected) else expected == value
        if not value_matches:
            raise ExcelValidationError(
                f"expected value '{expected}', found '{value}'", where=cell
            )

    def walk_schema(self, schema, visitor):
        """
        Walk an arbitrary schema definition, calling a visitor method for
        each SchemaCell we find in the structure.

        In the context of our spreadsheet library, a schema definition is
        simply an arbitrary Python data structure -- either a dict, or a
        list, or an arbitrary nesting of these -- where some of the values in
        a dictionary, or some of the items in a list, are derived from
        SchemaCell. In the cases where we find a SchemaCell, we invoke the
        visitor to produce a value based on it; in all other cases, we simply
        maintain the structure as-is.

        For instance:

            walk({"hello": "world", "goodbye": SchemaCell(...)}, get_value)
                -->
            {"hello": "world", "goodbye": 42}
        """
        if isinstance(schema, dict):
            result = {k: self.walk_schema(v, visitor) for k, v in schema.items()}
        elif isinstance(schema, list):
            result = [self.walk_schema(item, visitor) for item in schema]
        elif is_schema_cell(schema):
            result = visitor(schema)
        else:
            # Schema is a raw value; pass it through.
            result = schema
        return result

    def schema(self, schema, location=None, sheet=None, col=None, row=None):
        """
        Return values in a structured form, replacing all SchemaCells found
        in a schema with actual data from the spreadsheet.
        """
        sheet, col, row = parse_location_or_default(location, sheet, col, row)

        def _visitor(schema_cell):
            return self.value(schema_cell, sheet=sheet, col=col, row=row)

        return self.walk_schema(schema, _visitor)

    def schema_list(
        self,
        schema,
        locations=None,
        start=None,
        end=None,
        location=None,
        sheet=None,
        col=None,
        row=None,
    ):
        """
        Repeatedly call schema() with a varying location value each time.
        Return a list of the results.

        There are two ways to specify the varying locations. Either pass
        in a iterable for `locations` -- this can either be (sheet, col, row) tuples
        *or* naked row/column numbers, or provide a `start` or `end` which will
        be tossed to `location_range(...)` (see rowcol.py for details).

        In addition, default values for `location` --> (`sheet`, `col`, `row`)
        can be provided if the locations themselves are not complete.
        """
        # Determine the default values, if any
        sheet, col, row = parse_location_or_default(location, sheet, col, row)

        # Construct the varying locations
        locations = locations or location_range(start, end)

        def _schema(location):
            """Call schema on a location, defaulting to default values if needed."""
            sheet_, col_, row_ = parse_location_or_default(location, sheet, col, row)
            return self.schema(schema, sheet=sheet_, col=col_, row=row_)

        return [_schema(location) for location in locations]

    def schema_rect(
        self, schema, locations=None, location=None, sheet=None, col=None, row=None
    ):
        """
        Repeatedly call schema() with a varying location value each time.
        Return a two-dimensional list of the results.

        To specify location, pass in an iterable of iterables for `locations`;
        each should be in the form (sheet, col, row).

        In addition, default values for `location` --> (`sheet`, `col`, `row`)
        can be provided if the locations themselves are not complete.

        By default, we return a row-major array; this can be flipped.
        """
        # Determine the default values, if any
        sheet, col, row = parse_location_or_default(location, sheet, col, row)

        def _schema(location):
            """Call schema on a location, defaulting to default values if needed."""
            sheet_, col_, row_ = parse_location_or_default(location, sheet, col, row)
            return self.schema(schema, sheet=sheet_, col=col_, row=row_)

        return [[_schema(location) for location in inner] for inner in locations]

    def is_valid(self, ctx):
        """Validate the spreadsheet; return False if not possible."""
        try:
            self.clean(ctx)
        except ExcelError as e:
            self.errors.append(e)
        return len(self.errors) == 0

    def clean(self, ctx):
        """
        Validate the spreadsheet.

        Updates the underlying cleaned_data dictionary as appropriate.

        Adds ExcelValidationErrors to the errors list when encountering
        validation errors that do not impact the ability to further validate;
        raises ExcelValidationError when encountering a hard-stop validation
        error.
        """
        pass
