import os
import requests
import xml.etree.ElementTree as ET

import click

from openpyxl import load_workbook
from remark.lib.memoizer import delay_file_memoize
from remark.geo.usa_census import get_usa_census_data

FREE_MAP_TOOLS_URL = "https://www.freemaptools.com/ajax/us/get-all-zip-codes-inside.php"
FREE_MAP_REFERER = "https://www.freemaptools.com/find-zip-codes-inside-radius.htm"

USER_AGENT = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_12_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/72.0.3626.121 Safari/537.36"
MILES_KILOMETERS_RATIO = 1.60934

CACHE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "../../data/cache")
DEFAULT_TEMPLATE_PATH = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "../templates/tam-template.xlsx"
)
DEFAULT_OUTFILE_PATH = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "../../dist/remarkably-tam-test.xlsx"
)


VERBOSE = False


def print_verbose(*args, **kwargs):
    if VERBOSE:
        print(*args, **kwargs)


@delay_file_memoize(cache_dir=CACHE_DIR)
def fetch_zip_codes(latitude, longitude, radius):
    params = {
        "radius": radius,
        "lat": latitude,
        "lng": longitude,
        "rn": 8192,
        "showPOboxes": "true",
    }
    headers = {"user-agent": USER_AGENT, "referer": FREE_MAP_REFERER}
    response = requests.get(FREE_MAP_TOOLS_URL, params=params, headers=headers)
    root = ET.fromstring(response.text)
    result = []
    for child in root:
        result.append(child.attrib["postcode"])
    return result


def write_labeled_data(ws, title, labels, data, start_row):
    if len(data) != len(labels):
        print_verbose(labels)
        print_verbose(data)
        raise Exception(f"{title} Data length and label length is not equal.")

    ws.cell(column=1, row=start_row, value=title)
    for x in range(len(data)):
        actual_row = x + 1 + start_row
        ws.cell(column=1, row=actual_row, value=labels[x])
        ws.cell(column=2, row=actual_row, value=data[x])
    return None


def write_label_item(ws, label, item, row):
    ws.cell(column=1, row=row, value=label)
    ws.cell(column=2, row=row, value=item)


AGE_SEGMENT_LABELS = [
    "85+",
    "80-84",
    "75-79",
    "70-74",
    "67-69",
    "65-66",
    "62-64",
    "60-61",
    "55-59",
    "50-54",
    "45-49",
    "40-44",
    "35-39",
    "30-34",
    "25-29",
    "22-24",
    "21",
    "20",
    "18-19",
    "15-17",
    "10-14",
    "5-9",
    "0-4",
]

HOUSEHOLD_TYPE_LABELS = [
    "Married",
    "Single Female",
    "Single Male",
    "One Person",
    "Non-Family",
]

INCOME_DIST_LABELS = [
    "$200k",
    "$150-200k",
    "$125-150k",
    "$100-125k",
    "$75-100k",
    "$60-75k",
    "$50-60k",
    "$45-50k",
    "$40-45k",
    "$35-40k",
    "$30-35k",
    "$25-30k",
    "$20-25k",
    "$15-20k",
    "$10-15k",
    "< $10k",
]
AGE_SEGMENTS = ["18-24", "25-34", "35-44", "45-54", "55-64", "65+"]

ZIP_DATA_SHEET_NAME = "Zip Data {}"


def fill_zip_worksheet(workbook, zipcode):
    # Fetch all data first - if a Zip Code doesn't work. Ignore it.
    census_data = get_usa_census_data(zipcode)

    print_verbose("POP")
    print_verbose(census_data)

    # Create spreadsheet
    worksheet = workbook.create_sheet(title=ZIP_DATA_SHEET_NAME.format(zipcode))
    write_label_item(worksheet, "Total Population", census_data.population, 1)
    write_label_item(worksheet, "Number of Households", census_data.number_of_households, 2)
    if census_data.has_data:
        age_segments = census_data.age_segments
        household_type = census_data.households_by_type
        income_dist = census_data.income_distributions
    else:
        age_segments = [0.0 for _ in range(len(AGE_SEGMENT_LABELS))]
        household_type = [0.0 for _ in range(len(HOUSEHOLD_TYPE_LABELS))]
        income_dist = [0.0 for _ in range(len(INCOME_DIST_LABELS))]

    write_labeled_data(
        worksheet, "Population by Age Segment", AGE_SEGMENT_LABELS, age_segments, 4
    )
    write_labeled_data(
        worksheet,
        "Household by Type",
        HOUSEHOLD_TYPE_LABELS,
        household_type,
        len(age_segments) + 6,
    )
    write_labeled_data(
        worksheet,
        "Income Distribution",
        INCOME_DIST_LABELS,
        income_dist,
        len(age_segments) + len(household_type) + 8,
    )


def pop_formula(zip_codes, percent_values):
    formula = []
    for zip_code in zip_codes:
        subformula = []
        tab_name = ZIP_DATA_SHEET_NAME.format(zip_code)
        for val in percent_values:
            subformula.append(f"'{tab_name}'!B{val}")
        sf = f"('{tab_name}'!B1*({'+'.join(subformula)}))"
        formula.append(sf)
    result = "+".join(formula)
    return f"=ROUND({result}, 0)"


INCOME_DIST_LEVELS = (
    (0, 52),
    (10000, 51),
    (15000, 50),
    (20000, 49),
    (25000, 48),
    (30000, 47),
    (35000, 46),
    (40000, 45),
    (45000, 44),
    (50000, 43),
    (60000, 42),
    (75000, 41),
    (100000, 40),
    (125000, 39),
    (150000, 38),
    (200000, 37),
)


def fill_computation_tab(worksheet, zip_codes, income_groups):
    # Add ACS Population Reference Sumed across all zip_codes
    age_group_pop_schema = (
        (3, ("23", "22", "21", "20")),  # 18-24
        (4, ("18", "19")),  # 25-34
        (5, ("17", "16")),  # 35-44
        (6, ("15", "14")),  # 45-54
        (7, ("13", "12", "11")),  # 55-64
        (8, ("10", "9", "8", "7", "6", "5")),  # 65+
    )
    for dest_row, percent_values in age_group_pop_schema:
        formula = pop_formula(zip_codes, percent_values)
        worksheet.cell(column=2, row=dest_row, value=formula)

    start_x = 0
    for y in range(len(income_groups)):
        income_group = income_groups[y]
        if len(income_groups) - 1 == y:
            next_income_group = None
        else:
            next_income_group = income_groups[y + 1]
        numerator = []
        denominator = []
        for x in range(start_x, len(INCOME_DIST_LEVELS)):
            if (
                next_income_group is not None
                and next_income_group <= INCOME_DIST_LEVELS[x][0]
            ):
                print_verbose(
                    f"Income Break: {income_group} <= {INCOME_DIST_LEVELS[x][0]}"
                )
                break
            if income_group <= INCOME_DIST_LEVELS[x][0]:
                for zip_code in zip_codes:
                    tab_name = ZIP_DATA_SHEET_NAME.format(zip_code)
                    numerator.append(
                        f"('{tab_name}'!B{INCOME_DIST_LEVELS[x][1]}*'{tab_name}'!B1)"
                    )
            start_x = x + 1

        for zip_code in zip_codes:
            tab_name = ZIP_DATA_SHEET_NAME.format(zip_code)
            denominator.append(f"'{tab_name}'!B1")

        numerator_str = f"=({'+'.join(numerator)})"
        denominator_str = f"=({'+'.join(denominator)})"
        # final_formula = f"=({'+'.join(numerator)})/({'+'.join(denominator)})"
        worksheet.cell(column=1, row=20 + y, value=income_groups[y])
        # worksheet.cell(column=2, row=20 + y, value=final_formula)
        worksheet.cell(column=3, row=20 + y, value=numerator_str)
        worksheet.cell(column=4, row=20 + y, value=denominator_str)

    # Total population
    formula = []
    for zip_code in zip_codes:
        tab_name = ZIP_DATA_SHEET_NAME.format(zip_code)
        formula.append(f"'{tab_name}'!B1")
    worksheet.cell(column=2, row=29, value=f"={'+'.join(formula)}")

    # Add Household Types
    numerator = []
    for zip_code in zip_codes:
        tab_name = ZIP_DATA_SHEET_NAME.format(zip_code)
        numerator.append(
            f"(('{tab_name}'!B30+'{tab_name}'!B31+'{tab_name}'!B32)*'{tab_name}'!B1)"
        )
    final_formula = f"=({'+'.join(numerator)})/'Computation'!B29"
    worksheet.cell(column=2, row=25, value=final_formula)


def write_labeled_property(
    worksheet, search_value, write_value, search_column=1, write_column=2
):
    for x in range(1, 500):
        cell = worksheet.cell(column=search_column, row=x)
        if cell.value == search_value:
            return worksheet.cell(column=write_column, row=x, value=write_value)
    raise Exception("Did not find required label for value.")


def write_rti_info(worksheet, rti_income_groups, rti_rental_rates, rti_target):
    for x in range(len(rti_income_groups)):
        worksheet.cell(column=2 + x, row=26, value=rti_income_groups[x])
    for y in range(len(rti_rental_rates)):
        worksheet.cell(column=1, row=27 + y, value=rti_rental_rates[y])
    write_labeled_property(worksheet, "Target RTI", rti_target)


def write_tam_type(worksheet, lat, lon, tam_type, tam_data):
    if tam_type not in ["zipcodes", "radius"]:
        raise Exception("Tam Type is invalid")
    worksheet.cell(column=2, row=4, value=tam_type)
    if tam_type == "zipcodes":
        for x in range(len(tam_data)):
            worksheet.cell(column=3 + x, row=4, value=tam_data[x])
    elif tam_type == "radius":
        worksheet.cell(column=3, row=4, value=tam_data)
        worksheet.cell(column=4, row=4, value="mi")

    write_labeled_property(worksheet, "Coordinates", lat)
    write_labeled_property(worksheet, "Coordinates", lon, write_column=3)


def write_output_properties(worksheet, loc, max_rent, avg_rent, min_rent, usvs):
    write_labeled_property(worksheet, "City, State", loc)
    write_labeled_property(worksheet, "Maximum Rent", max_rent)
    write_labeled_property(worksheet, "Average Rent", avg_rent)
    write_labeled_property(worksheet, "Minimum Rent", min_rent)
    # Age Segment USVS
    for x in range(len(AGE_SEGMENTS)):
        write_labeled_property(worksheet, AGE_SEGMENTS[x], usvs[x], write_column=3)


# Meridian
# Zip Codes: 84101,84111,84102,84112,84115,84105
# Income Groups: 75000, 50000, 35000

# El Cortez
# Zip Codes: 85013
# Income Groups: 25000, 45000, 60000


def build_tam_data_for_zip_codes(workbook, zip_codes, income_groups):
    # generate Zip Code Worksheet
    live_zipcodes = []
    for zip_code in zip_codes:
        try:
            print_verbose(f"Starting zip: {zip_code}")
            fill_zip_worksheet(workbook, zip_code)
            live_zipcodes.append(zip_code)
            print_verbose(f"Completed zip: {zip_code}")
        except Exception as e:
            print_verbose(f"FAILED zip: {zip_code}")
            raise e

    fill_computation_tab(workbook["Computation"], live_zipcodes, income_groups)

    # TODO what's all this, then?
    # fetch_zip_codes(latitude, longitude, radius * MILES_KILOMETERS_RATIO)
    # fetch_age_segments_by_zip('85013')
    # parse_age_segments(STAT_ATLAS_AGE_CONTENT)
    # fetch_household_income_distribution('85013')


def build_tam_data_for_location(workbook, location, radius, income_groups):
    zip_codes = fetch_zip_codes(
        location[0], location[1], radius * MILES_KILOMETERS_RATIO
    )
    print_verbose(f"zipcodes: {zip_codes}")
    return build_tam_data_for_zip_codes(workbook, zip_codes, income_groups)


def build_tam_data(
    zip_codes,
    lat,
    lon,
    loc,
    radius,
    income_groups,
    rti_income_groups,
    rti_rental_rates,
    rti_target,
    age,
    max_rent,
    avg_rent,
    min_rent,
    usvs,
    templatefile=DEFAULT_TEMPLATE_PATH,
):

    # Must have 4+ rental rates
    if len(rti_rental_rates) < 4:
        raise Exception("Must have 4 or more RTI rental rates")

    if len(rti_income_groups) < 4:
        raise Exception("Must have 4 or more RTI income groups")

    if len(usvs) != 6:
        raise Exception("There must be six USV entries")

    if rti_target is None:
        rti_target = 0.3333

    # Load workbook
    workbook = load_workbook(templatefile)

    # Delete Sample Data tab
    workbook.remove(workbook["Sample Zip Data"])

    # Build the data into the workbook
    if zip_codes:
        build_tam_data_for_zip_codes(workbook, zip_codes, income_groups)
        write_tam_type(workbook["Output"], lat, lon, "zipcodes", zip_codes)
    elif lat and lon and radius:
        build_tam_data_for_location(workbook, (lat, lon), radius, income_groups)
        write_tam_type(workbook["Output"], lat, lon, "radius", radius)
    else:
        raise click.UsageError(
            "You must specify either zip_codes *or* a location and radius."
        )

    # Write Output tab Data
    write_rti_info(workbook["Output"], rti_income_groups, rti_rental_rates, rti_target)
    write_output_properties(workbook["Output"], loc, max_rent, avg_rent, min_rent, usvs)

    return workbook


@click.command()
@click.option("--lat", type=float, required=True, help="The latitude.")
@click.option("--lon", type=float, required=True, help="The longitude.")
@click.option(
    "--loc",
    type=str,
    required=True,
    help="The City and State of the location in {city},{state} format",
)
@click.option("-r", "--radius", type=float, help="A radius in miles.")
@click.option(
    "-z", "--zip", "zip_codes", type=int, multiple=True, help="A list of ZIP codes."
)
@click.option(
    "-i",
    "--income",
    "income_groups",
    type=int,
    multiple=True,
    required=True,
    help="A list of income group codes.",
)
@click.option(
    "-g",
    "--rti-income",
    "rti_income_groups",
    type=int,
    multiple=True,
    required=True,
    help="A list of RTI income limits",
)
@click.option(
    "-m",
    "--rti-rent",
    "rti_rental_rates",
    type=int,
    multiple=True,
    required=True,
    help="A list of RTI rental rates",
)
@click.option(
    "-e", "--rti-target", "rti_target", type=float, help="RTI target percent."
)
@click.option("-a", "--age", type=int, required=True, help="Average Tenant Age")
@click.option(
    "-b", "--max-rent", "max_rent", type=int, required=True, help="Maximum Rent"
)
@click.option(
    "-c", "--avg-rent", "avg_rent", type=int, required=True, help="Average Rent"
)
@click.option(
    "-d", "--min-rent", "min_rent", type=int, required=True, help="Minimum Rent"
)
@click.option(
    "-u",
    "--usvs",
    type=int,
    multiple=True,
    required=True,
    help="Unique Site Visitors. There must be 6 entries.",
)
@click.option(
    "-t",
    "--templatefile",
    default=DEFAULT_TEMPLATE_PATH,
    type=click.Path(exists=True),
    help="The TAM XLS template path.",
)
@click.option(
    "-o",
    "--outfile",
    default=DEFAULT_OUTFILE_PATH,
    type=click.Path(),
    help="The output XLS filename",
)
def main(
    zip_codes,
    lat,
    lon,
    loc,
    radius,
    income_groups,
    rti_income_groups,
    rti_rental_rates,
    rti_target,
    age,
    max_rent,
    avg_rent,
    min_rent,
    usvs,
    templatefile,
    outfile,
):
    workbook = build_tam_data(
        zip_codes,
        lat,
        lon,
        loc,
        radius,
        income_groups,
        rti_income_groups,
        rti_rental_rates,
        rti_target,
        age,
        max_rent,
        avg_rent,
        min_rent,
        usvs,
        templatefile,
    )
    # Save the resulting workbook
    workbook.save(filename=outfile)


if __name__ == "__main__":
    VERBOSE = True
    build_tam_data()
