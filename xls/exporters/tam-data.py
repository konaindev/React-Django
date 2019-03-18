import sys
import os
import pickle
import requests
import time
import xml.etree.ElementTree as ET
from bs4 import BeautifulSoup
from openpyxl import load_workbook

FREE_MAP_TOOLS_URL = "https://www.freemaptools.com/ajax/us/get-all-zip-codes-inside.php"
FREE_MAP_REFERER = "https://www.freemaptools.com/find-zip-codes-inside-radius.htm"

STAT_ATLAS_AGE_URL = "https://statisticalatlas.com/zip/%s/Age-and-Sex"
STAT_ATLAS_HOUSEHOLD_URL = "https://statisticalatlas.com/zip/%s/Household-Types"
STAT_ATLAS_HOUSEHOLD_INCOME_URL = "https://statisticalatlas.com/zip/%s/Household-Income"
STAT_ATLAS_REFER = "https://statisticalatlas.com/United-States/Overview"

USER_AGENT = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_12_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/72.0.3626.121 Safari/537.36"
MILES_KILOMETERS_RATIO = 1.60934


def memoize(label):
    def outer_cache(func):
        def inner_cache(*args, **kwargs):
            filename = label
            for arg in args:
                filename += ":%s" % str(arg)
            filename += ".pkl"
            filepath = "../../data/zipcode/%s" % filename
            if os.path.exists(filepath):
                print("Read from disk: %s" % filepath)
                file_ref = open(filepath, "rb")
                return pickle.load(file_ref)
            else:
                print("Writing to disk: %s" % filepath)
                result = func(*args)
                file_ref = open(filepath, "wb")
                pickle.dump(result, file_ref)
                time.sleep(20)
                print("Finished writing.")
                return result

        return inner_cache

    return outer_cache


@memoize("fetch-zip-codes")
def fetch_zip_codes(latitude, longitude, radius):
    params = {
        "radius": radius,
        "lat": latitude,
        "lng": longitude,
        "rn": 8192,
        "showPOboxes": "true",
    }
    headers = {"user-agent": USER_AGENT, "referer": FREE_MAP_REFERER}
    response = requests.get(FREE_MAP_TOOLS_URL, params=params, headers=headers)
    root = ET.fromstring(response.text)
    result = []
    for child in root:
        result.append(child.attrib["postcode"])
    return result


def find_population(el):
    return el.has_attr("title") and el["title"] == "Population"


def find_households(el):
    return el.has_attr("title") and el["title"] == "Households"


@memoize("fetch-population")
def fetch_population(zipcode):
    url = STAT_ATLAS_AGE_URL % zipcode
    headers = {"user-agent": USER_AGENT, "referer": STAT_ATLAS_REFER}
    response = requests.get(url, headers=headers)
    soup = BeautifulSoup(response.text, features="html.parser")
    pop_th = soup.find_all(find_population)[0]
    td_value = str(pop_th.td.text)
    population = int(td_value.replace(",", ""))
    print("Population: %d" % population)

    house_th = soup.find_all(find_households)[0]
    td_value = str(house_th.td.text)
    households = int(td_value.replace(",", ""))
    print("households: %d" % population)

    return (population, households)


def fetch_svg(base_url, zipcode, figure_id):
    def find_figure(el):
        if el.has_attr("id") and el["id"] == figure_id:
            return True
        return False

    url = base_url % zipcode
    headers = {"user-agent": USER_AGENT, "referer": STAT_ATLAS_REFER}
    response = requests.get(url, headers=headers)

    soup = BeautifulSoup(response.text, features="html.parser")
    figures = soup.find_all(find_figure)
    if len(figures) == 0:
        ref = 'id="%s"' % figure_id
        print("Searching for %s..." % ref)
        print("Results: %d" % (response.text.find(ref)))
        raise Exception("Could not find the following figure: %s" % figure_id)

    try:
        svg = figures[0].find_all("svg")[0]
    except Exception as e:
        print("Exception: %s" % str(e))
        print("Length: %d" % len(response.text))
        print(figures[0].find("svg"))
        raise e
    return svg


@memoize("fetch-age-segments")
def fetch_age_segements_by_zip(zipcode):
    svg = fetch_svg(STAT_ATLAS_AGE_URL, zipcode, "figure/age-structure")
    gs = svg.g.find_all("g")
    result = []
    for x in range(len(gs)):
        # print("%d: %s" % (x, gs[x]))
        if x >= 26 and x < 49:
            txt = gs[x].title.text
            value = float(txt.replace("%", ""))
            result.append(value / 100.0)
    # print(result)
    return result


@memoize("fetch-household-type")
def fetch_household_type(zipcode):
    svg = fetch_svg(STAT_ATLAS_HOUSEHOLD_URL, zipcode, "figure/household-types")
    gs = svg.g.find_all("g")
    result = []
    for x in range(len(gs)):
        # print("%d: %s" % (x, gs[x]))
        if x >= 7:
            txt = gs[x].title.text
            value = float(txt.replace("%", ""))
            result.append(value / 100.0)
    return result


@memoize("fetch-household-income")
def fetch_household_income(zipcode):
    svg = fetch_svg(
        STAT_ATLAS_HOUSEHOLD_INCOME_URL, zipcode, "figure/household-income-percentiles"
    )
    gs = svg.g.find_all("g")
    result = [gs[8], gs[10], gs[12], gs[14], gs[16], gs[18]]
    for x in range(len(result)):
        txt = result[x].title.text
        result[x] = float(txt.replace("$", "").replace(",", ""))
    return result


@memoize("fetch-household-income-dist")
def fetch_household_income_distribution(zipcode):
    svg = fetch_svg(
        STAT_ATLAS_HOUSEHOLD_INCOME_URL, zipcode, "figure/household-income-distribution"
    )
    print(svg)
    gs = svg.g.find_all("g")
    result = []
    for x in range(len(gs)):
        if x >= 19 and x < 35:
            txt = gs[x].title.text
            value = float(txt.replace("%", ""))
            result.append(value / 100.0)
    print(result)
    return result


def write_labeled_data(ws, title, labels, data, start_row):
    ws.cell(column=1, row=start_row, value=title)
    for x in range(len(data)):
        actual_row = x + 1 + start_row
        ws.cell(column=1, row=actual_row, value=labels[x])
        ws.cell(column=2, row=actual_row, value=data[x])
    return None


def write_label_item(ws, label, item, row):
    ws.cell(column=1, row=row, value=label)
    ws.cell(column=2, row=row, value=item)


AGE_SEGMENT_LABELS = [
    "85+",
    "80-84",
    "75-79",
    "70-74",
    "67-69",
    "65-66",
    "62-64",
    "60-61",
    "55-59",
    "50-54",
    "45-49",
    "40-44",
    "35-39",
    "30-34",
    "25-29",
    "22-24",
    "21",
    "20",
    "18-19",
    "15-17",
    "10-14",
    "5-9",
    "0-4",
]

HOUSEHOLD_TYPE_LABELS = [
    "Married",
    "Single Female",
    "Single Male",
    "One Person",
    "Non-Family",
]

INCOME_DIST_LABELS = [
    "$200k",
    "$150-200k",
    "$125-150k",
    "$100-125k",
    "$75-100k",
    "$60-75k",
    "$50-60k",
    "$45-50k",
    "$40-45k",
    "$35-40k",
    "$30-35k",
    "$25-30k",
    "$20-25k",
    "$15-20k",
    "$10-15k",
    "< $10k",
]

ZIP_DATA_SHEET_NAME = "Zip Data %s"


def fill_zip_worksheet(workbook, zipcode):
    # Fetch all data first - if a Zip Code doesn't work. Ignore it.
    population = fetch_population(zipcode)
    age_segments = fetch_age_segements_by_zip(zipcode)
    household_type = fetch_household_type(zipcode)
    # household_income = fetch_household_income(zipcode)
    income_dist = fetch_household_income_distribution(zipcode)

    print("POP")
    print(population)

    # Create speadsheet
    worksheet = workbook.create_sheet(title=ZIP_DATA_SHEET_NAME % zipcode)
    write_label_item(worksheet, "Total Population", population[0], 1)
    write_label_item(worksheet, "Number of Households", population[1], 2)
    write_labeled_data(
        worksheet, "Population by Age Segment", AGE_SEGMENT_LABELS, age_segments, 4
    )
    write_labeled_data(
        worksheet,
        "Household by Type",
        HOUSEHOLD_TYPE_LABELS,
        household_type,
        len(age_segments) + 6,
    )
    write_labeled_data(
        worksheet,
        "Income Distribution",
        INCOME_DIST_LABELS,
        income_dist,
        len(age_segments) + len(household_type) + 8,
    )


def pop_formula(zipcodes, percent_values):
    formula = []
    for zip in zipcodes:
        subformula = []
        tab_name = ZIP_DATA_SHEET_NAME % zip
        for val in percent_values:
            subformula.append("'%s'!B%s" % (tab_name, val))
        sf = "('%s'!B1*(%s))" % (tab_name, "+".join(subformula))
        formula.append(sf)
    result = "+".join(formula)
    return "=ROUND(%s, 0)" % result


INCOME_DIST_LEVELS = (
    (200000, 37),
    (150000, 38),
    (125000, 39),
    (100000, 40),
    (75000, 41),
    (60000, 42),
    (50000, 43),
    (45000, 44),
    (40000, 45),
    (35000, 46),
    (30000, 47),
    (25000, 48),
    (20000, 49),
    (15000, 50),
    (10000, 51),
    (0, 52),
)


def fill_computation_tab(worksheet, zipcodes):
    # Add ACS Population Reference Sumed across all zipcodes
    age_group_pop_schema = (
        (3, ("23", "22", "21", "20")),  # 18-24
        (4, ("18", "19")),  # 25-34
        (5, ("17", "16")),  # 35-44
        (6, ("15", "14")),  # 45-54
        (7, ("13", "12", "11")),  # 55-64
        (8, ("10", "9", "8", "7", "6", "5")),  # 65+
    )
    for dest_row, percent_values in age_group_pop_schema:
        formula = pop_formula(zipcodes, percent_values)
        worksheet.cell(column=2, row=dest_row, value=formula)

    cell_formulas = []
    start_x = 0
    for income_group in INCOME_GROUPS:
        numerator = []
        denominator = []
        for x in range(start_x, len(INCOME_DIST_LEVELS)):
            for zip in zipcodes:
                tab_name = ZIP_DATA_SHEET_NAME % zip
                numerator.append(
                    "('%s'!B%d*'%s'!B1)"
                    % (tab_name, INCOME_DIST_LEVELS[x][1], tab_name)
                )

            denominator.append("'%s'!B1" % tab_name)

            if income_group == INCOME_DIST_LEVELS[x][0]:
                start_x = x + 1
                break
        final_formula = "=(%s)/(%s)" % ("+".join(numerator), "+".join(denominator))
        cell_formulas.append(final_formula)

    for x in range(3):
        worksheet.cell(column=1, row=29 + x, value=INCOME_GROUPS[x])
        worksheet.cell(column=2, row=29 + x, value=cell_formulas[x])

    # Total population
    formula = []
    for zip in zipcodes:
        tab_name = ZIP_DATA_SHEET_NAME % zip
        formula.append("'%s'!B1" % tab_name)
    worksheet.cell(column=2, row=38, value="=%s" % "+".join(formula))

    # Add Household Types
    numerator = []
    for zip in zipcodes:
        tab_name = ZIP_DATA_SHEET_NAME % zip
        numerator.append(
            "(('%s'!B%d+'%s'!B%d+'%s'!B%d)*'%s'!B1)"
            % (tab_name, 30, tab_name, 31, tab_name, 32, tab_name)
        )
    final_formula = "=(%s)/'Computation'!B38" % "+".join(numerator)
    worksheet.cell(column=2, row=34, value=final_formula)


ZIP_CODES = [85013]

# Meridian
# Zip Codes: 84101,84111,84102,84112,84115,84105
# Income Groups: 75000, 50000, 35000

# El Cortez
# Zip Codes: 85013
# Income Groups: 25000, 45000, 60000

INCOME_GROUPS = [25000, 45000, 60000]


def main():
    # latitude = sys.argv[1]
    # longitude = sys.argv[2]
    # radius = float(sys.argv[3])

    zipcodes = ZIP_CODES  # fetch_zip_codes(latitude, longitude, radius)
    workbook = load_workbook("../templates/tam-template.xlsx")

    # generate Zip Code Worksheet
    for zipcode in zipcodes:
        try:
            print("Starting ZipCode: %s" % zipcode)
            fill_zip_worksheet(workbook, zipcode)
            print("Completed ZipCode: %s" % zipcode)
        except Exception as e:
            print("ZipCode Failed: %s" % zipcode)
            raise e

    fill_computation_tab(workbook["Computation"], zipcodes)

    # delete sample data tab
    workbook.remove(workbook["Sample Zip Data"])

    workbook.save(filename="../../dist/remarkably-tam-test.xlsx")


# fetch_zip_codes(latitude, longitude, radius * MILES_KILOMETERS_RATIO)
# fetch_age_segements_by_zip('85013')
# parse_age_segments(STAT_ATLAS_AGE_CONTENT)
# fetch_household_income_distribution('85013')

main()
